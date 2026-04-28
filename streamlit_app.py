import os
import math
import importlib
import json
import re
import zipfile
from xml.etree import ElementTree as ET
from datetime import date
from datetime import timedelta
from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import excel_analysis

st.set_page_config(
    page_title="S관 생산 필요수량 대시보드",
    layout="wide",
    initial_sidebar_state="collapsed",
)


DATA_DIR = "data"
KST = ZoneInfo("Asia/Seoul")
REPO_EXCEL_CANDIDATES = [
    "s관 부족수량.xlsx",
    os.path.join(DATA_DIR, "s관 부족수량.xlsx"),
]
TEMPLATE_XLSX_PATH = "업로드 양식.xlsx"
OUT_DIR = "out"
STREAMLIT_CONFIG_PATH = os.path.join(".streamlit", "config.toml")
DASHBOARD_LINKS_PATH = "dashboard_links.json"

# Risk tab business defaults (fixed; no user settings)
RISK_CAPA_RUN_DAYS = 7
RISK_YELLOW_BUFFER_DAYS = 1.0
# Fixed WIP/queue lead time per process (calendar days)
RISK_WIP_DAYS_PER_PROCESS = 1.0
# Scheduling start offset (calendar days). Use 0 to allow same-day completion dates like 4/22.
RISK_SCHED_START_OFFSET_DAYS = 0.0


@st.cache_data(show_spinner=False)
def _load_theme_from_config_cached(_mtime: float) -> dict:
    _ = _mtime  # cache-buster when file changes
    try:
        if not os.path.exists(STREAMLIT_CONFIG_PATH):
            return {}
        with open(STREAMLIT_CONFIG_PATH, "rb") as f:
            import tomllib  # py3.11+

            data = tomllib.load(f)
        return data.get("theme", {}) if isinstance(data, dict) else {}
    except Exception:
        return {}


def _today_kst() -> date:
    return datetime.now(tz=KST).date()


def _end_of_month(d: date) -> date:
    if d.month == 12:
        first_next = date(d.year + 1, 1, 1)
    else:
        first_next = date(d.year, d.month + 1, 1)
    return first_next - timedelta(days=1)


def _load_dashboard_links(path: str = DASHBOARD_LINKS_PATH) -> list[dict[str, str]]:
    """
    Load external dashboard links from a local json file.

    Recommended schema:
      [{"label": "...", "url": "https://..."}]

    Also accepts:
      {"links": [...]} and "name" instead of "label".
    """
    try:
        if not os.path.exists(path):
            return []
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict) and isinstance(data.get("links"), list):
            data = data["links"]
        if not isinstance(data, list):
            return []
        out: list[dict[str, str]] = []
        for item in data:
            if not isinstance(item, dict):
                continue
            label = str(item.get("label") or item.get("name") or "").strip()
            url = str(item.get("url") or "").strip()
            if label and url:
                out.append({"label": label, "url": url})
        return out
    except Exception:
        return []


def _table_height_for_rows(
    n_rows: int,
    *,
    min_height: int,
    max_height: int,
    header_px: int = 110,
    row_px: int = 34,
) -> int:
    n = max(0, int(n_rows))
    h = int(header_px + (n * row_px))
    return max(min_height, min(max_height, h))


def _pre_widget_single_select_fix(*, key: str, default: str, options: list[str]) -> None:
    """
    Safe to call BEFORE the widget is instantiated in the current run.
    Fixes invalid/cleared value in session_state so the widget shows `default`.
    """
    v = st.session_state.get(key)
    if isinstance(v, list):
        v = v[0] if v else None
    if isinstance(v, str):
        v = v.strip()
    if v not in options:
        st.session_state[key] = default


def _on_change_single_select(key: str, default: str, options: list[str]) -> None:
    """
    Callback: safe to mutate session_state for the widget key.
    Used to snap back to default when user clears the selection.
    """
    v = st.session_state.get(key)
    if isinstance(v, list):
        v = v[0] if v else None
    if isinstance(v, str):
        v = v.strip()
    if v not in options:
        st.session_state[key] = default


def _on_change_risk_grade_pills(*, key: str, grade_options: list[str]) -> None:
    """
    Risk tab grade filter callback.
    Selecting '해제' sets selection to all grades.
    """
    v = st.session_state.get(key)
    if not isinstance(v, list):
        v = [v] if v else []
    v = [str(x).strip() for x in v if str(x).strip()]
    st.session_state[key] = [g for g in v if g in grade_options]


def _coerce_single_value(value: str | None, *, default: str, options: list[str]) -> str:
    v = (value or "").strip()
    return v if v in options else default


def _pre_widget_multi_select_fix(*, key: str, default: list[str], options: list[str]) -> None:
    """
    Safe to call BEFORE the widget is instantiated in the current run.
    Fixes invalid/cleared multi-select value in session_state so the widget shows `default`.
    """
    v = st.session_state.get(key)
    if not isinstance(v, list):
        v = [v] if v else []
    v = [str(x).strip() for x in v if str(x).strip()]
    v = [x for x in v if x in options]
    if not v:
        st.session_state[key] = list(default)
    else:
        st.session_state[key] = v


def _on_change_multi_select_all_exclusive(key: str, all_value: str, options: list[str]) -> None:
    """
    Callback: multi-select where `all_value` behaves as exclusive default.
    - empty => [all_value]
    - selecting [all_value, ...] => remove `all_value` (keep specifics)
    """
    v = st.session_state.get(key)
    if not isinstance(v, list):
        v = [v] if v else []
    v = [str(x).strip() for x in v if str(x).strip()]
    v = [x for x in v if x in options]
    if not v:
        st.session_state[key] = [all_value]
        return
    if (all_value in v) and (len(v) > 1):
        v = [x for x in v if x != all_value]
        v = list(dict.fromkeys(v))
        st.session_state[key] = v if v else [all_value]
        return
    st.session_state[key] = v


def _coerce_multi_values(value: object, *, default: list[str], options: list[str]) -> list[str]:
    v = value
    if not isinstance(v, list):
        v = [v] if v else []
    out = [str(x).strip() for x in v if str(x).strip()]
    out = [x for x in out if x in options]
    return out if out else list(default)


def _find_repo_excel() -> str | None:
    for p in REPO_EXCEL_CANDIDATES:
        if os.path.exists(p):
            return p
    return None


def _file_mtime_label(path: str) -> str:
    try:
        def _xlsx_modified_ts(p: str) -> datetime | None:
            if not str(p).lower().endswith(".xlsx"):
                return None
            try:
                with zipfile.ZipFile(p) as zf:
                    core = zf.read("docProps/core.xml")
                root = ET.fromstring(core)
                ns = {
                    "dcterms": "http://purl.org/dc/terms/",
                }
                modified_el = root.find(".//dcterms:modified", ns)
                if modified_el is None or not (modified_el.text or "").strip():
                    return None
                raw = modified_el.text.strip()
                raw = raw.replace("Z", "+00:00")
                dt = datetime.fromisoformat(raw)
                return dt if dt.tzinfo is not None else dt.replace(tzinfo=ZoneInfo("UTC"))
            except Exception:
                return None

        # Prefer Excel's internal "modified" timestamp (prevents git checkout time confusion).
        ts = _xlsx_modified_ts(path)
        if ts is None:
            ts = datetime.fromtimestamp(os.path.getmtime(path), tz=KST)
        else:
            ts = ts.astimezone(KST)
        return ts.strftime("%Y-%m-%d %H:%M:%S %Z")
    except Exception:
        return "-"


def _read_bytes(path: str) -> bytes | None:
    try:
        with open(path, "rb") as f:
            return f.read()
    except Exception:
        return None


def _safe_mkdir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


@st.cache_data(show_spinner=False)
def _xlsx_sheet_names_cached(path: str, _mtime: float) -> set[str]:
    _ = _mtime  # cache-buster when file changes
    try:
        if not path or (not os.path.exists(path)):
            return set()
        with zipfile.ZipFile(path) as zf:
            wb = zf.read("xl/workbook.xml")
        root = ET.fromstring(wb)
        ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        sheets_el = root.find("m:sheets", ns)
        if sheets_el is None:
            return set()
        out: set[str] = set()
        for sh in sheets_el.findall("m:sheet", ns):
            name = (sh.attrib.get("name") or "").strip()
            if name:
                out.add(name)
        return out
    except Exception:
        # Fallback (should be rare): use pandas engine if workbook.xml is unavailable.
        try:
            xl = pd.ExcelFile(path)
            return set(str(x).strip() for x in xl.sheet_names if str(x).strip())
        except Exception:
            return set()


def _outputs_status(*, excel_path: str, out_dir: str) -> dict:
    """
    Fast up-to-date check for derived CSV outputs.

    Important UX: tab switches cause reruns; keep this check lightweight so the
    app feels instant when nothing changed.
    """
    due_csv = os.path.join(out_dir, "납기_제품군_공정별부족.csv")
    detail_csv = os.path.join(out_dir, "이니셜별_수주상세.csv")
    equip_code_target_csv = os.path.join(out_dir, "설비별_공정_제품코드_최소목표일.csv")
    prod_daily_csv = os.path.join(out_dir, "생산실적_공정별_일별양품.csv")

    try:
        excel_mtime = float(os.path.getmtime(excel_path))
    except Exception:
        return {"ok": False, "reason": "엑셀 파일을 찾을 수 없습니다."}

    # Base outputs must exist and be newer than the Excel file.
    base_ok = (
        os.path.exists(due_csv)
        and os.path.exists(detail_csv)
        and (os.path.getmtime(due_csv) >= excel_mtime)
        and (os.path.getmtime(detail_csv) >= excel_mtime)
    )
    if not base_ok:
        return {
            "ok": True,
            "needs_regen": True,
            "due_csv": due_csv,
            "detail_csv": detail_csv,
            "equip_code_target_csv": None,
            "prod_daily_csv": None,
        }

    # Optional outputs are required only if the corresponding sheets exist.
    sheet_names = _xlsx_sheet_names_cached(excel_path, excel_mtime)
    has_equip = "설비별" in sheet_names
    has_prod = "생산실적" in sheet_names

    equip_ok = os.path.exists(equip_code_target_csv) and (os.path.getmtime(equip_code_target_csv) >= excel_mtime)
    prod_ok = os.path.exists(prod_daily_csv) and (os.path.getmtime(prod_daily_csv) >= excel_mtime)

    if (has_equip and (not equip_ok)) or (has_prod and (not prod_ok)):
        return {
            "ok": True,
            "needs_regen": True,
            "due_csv": due_csv,
            "detail_csv": detail_csv,
            "equip_code_target_csv": equip_code_target_csv if has_equip else None,
            "prod_daily_csv": prod_daily_csv if has_prod else None,
        }

    return {
        "ok": True,
        "needs_regen": False,
        "regenerated": False,
        "due_csv": due_csv,
        "detail_csv": detail_csv,
        "equip_code_target_csv": equip_code_target_csv if (has_equip and equip_ok) else None,
        "prod_daily_csv": prod_daily_csv if (has_prod and prod_ok) else None,
    }


def _ensure_latest_outputs(*, excel_path: str, out_dir: str) -> dict:
    due_csv = os.path.join(out_dir, "납기_제품군_공정별부족.csv")
    detail_csv = os.path.join(out_dir, "이니셜별_수주상세.csv")
    equip_code_target_csv = os.path.join(out_dir, "설비별_공정_제품코드_최소목표일.csv")
    prod_daily_csv = os.path.join(out_dir, "생산실적_공정별_일별양품.csv")
    excel_mtime = os.path.getmtime(excel_path)

    if os.path.exists(due_csv) and os.path.exists(detail_csv):
        if os.path.getmtime(due_csv) >= excel_mtime and os.path.getmtime(detail_csv) >= excel_mtime:
            try:
                sheet_names = _xlsx_sheet_names_cached(excel_path, float(excel_mtime))
                has_equip = "설비별" in sheet_names
                has_prod = "생산실적" in sheet_names
            except Exception:
                has_equip = False
                has_prod = False

            equip_ok = os.path.exists(equip_code_target_csv) and os.path.getmtime(equip_code_target_csv) >= excel_mtime
            prod_ok = os.path.exists(prod_daily_csv) and os.path.getmtime(prod_daily_csv) >= excel_mtime

            # All required/available outputs exist and are up-to-date.
            if (not has_equip or equip_ok) and (not has_prod or prod_ok):
                return {
                    "ok": True,
                    "regenerated": False,
                    "needs_regen": False,
                    "due_csv": due_csv,
                    "detail_csv": detail_csv,
                    "equip_code_target_csv": equip_code_target_csv if has_equip and equip_ok else None,
                    "prod_daily_csv": prod_daily_csv if has_prod and prod_ok else None,
                }

    _safe_mkdir(out_dir)
    importlib.reload(excel_analysis)
    info = excel_analysis.export_due_process_shortage(file_path=excel_path, out_dir=out_dir)
    if not info.get("enabled"):
        return {"ok": False, "reason": info.get("reason") or "export failed"}
    # Optional: production actuals (for risk tab)
    try:
        prod_info = excel_analysis.export_production_daily_good_qty(file_path=excel_path, out_dir=out_dir)
        _ = prod_info
    except Exception:
        pass
    return {
        "ok": True,
        "regenerated": True,
        "needs_regen": False,
        "due_csv": due_csv,
        "detail_csv": detail_csv,
        "equip_code_target_csv": equip_code_target_csv if os.path.exists(equip_code_target_csv) else None,
        "prod_daily_csv": prod_daily_csv if os.path.exists(prod_daily_csv) else None,
    }


def _load_theme_from_config() -> dict:
    """
    Load Streamlit theme config (cached by file mtime).

    This is called frequently (table styling / CSS injection), so we cache to reduce
    redundant disk reads while keeping behavior identical when config changes.
    """

    try:
        mtime = os.path.getmtime(STREAMLIT_CONFIG_PATH) if os.path.exists(STREAMLIT_CONFIG_PATH) else -1.0
    except Exception:
        mtime = -1.0
    return _load_theme_from_config_cached(float(mtime))


def _style_dataframe_like_dashboard(df: pd.DataFrame) -> object:
    # Streamlit Styler rendering becomes noticeably slow on large frames.
    # Keep the dashboard snappy by skipping styling for big tables.
    try:
        if isinstance(df, pd.DataFrame):
            if (df.shape[0] > 1500) or (df.shape[1] > 60):
                return df
    except Exception:
        return df

    theme = _load_theme_from_config()
    bg = theme.get("backgroundColor", "#FBF7EE")
    sbg = theme.get("secondaryBackgroundColor", "#F2EBDD")
    text = theme.get("textColor", "#1B1B1B")
    try:
        return (
            df.style.set_properties(**{"background-color": bg, "color": text})
            .set_table_styles(
                [
                    {"selector": "th", "props": [("background-color", sbg), ("color", text)]},
                    {"selector": "td", "props": [("background-color", bg), ("color", text)]},
                ]
            )
        )
    except Exception:
        return df


def _apply_local_theme_css() -> None:
    theme = _load_theme_from_config()
    bg = theme.get("backgroundColor", "#FBF7EE")
    sbg = theme.get("secondaryBackgroundColor", "#F2EBDD")
    text = theme.get("textColor", "#1B1B1B")
    primary = theme.get("primaryColor", "#0A5C36")

    st.markdown(
        f"""
<style>
.stApp {{
  background-color: {bg} !important;
  color: {text} !important;
}}
[data-testid="stSidebar"] > div {{
  background-color: {sbg} !important;
}}
/* Make widgets/containers closer to local look across deployments */
div[data-testid="stVerticalBlockBorderWrapper"],
div[data-testid="stContainer"] {{
  background-color: transparent;
}}
a, a:visited {{
  color: {primary};
}}

/* Air / spacing */
.block-container {{
  padding-top: 2.2rem !important;
  padding-bottom: 2.2rem !important;
}}
div[data-testid="stVerticalBlock"] > div {{
  row-gap: 0.9rem;
}}
/* Pills & segmented controls spacing */
div[data-testid="stPills"] > div {{
  margin-top: 0.2rem;
  margin-bottom: 0.7rem;
}}
div[data-testid="stSegmentedControl"] > div {{
  margin-top: 0.2rem;
  margin-bottom: 0.7rem;
}}

/* Sidebar layout */
.sb-title {{
  font-size: 17px;
  font-weight: 800;
  margin: 0.15rem 0 0.65rem 0;
}}
.sb-hr {{
  border: 0;
  border-top: 1px solid rgba(0, 0, 0, 0.08);
  margin: 0.8rem 0;
}}
.sb-kv {{
  border: 1px solid rgba(0, 0, 0, 0.08);
  border-radius: 12px;
  padding: 0.9rem 0.95rem;
  background: rgba(255, 255, 255, 0.55);
}}
.sb-kv .row {{
  display: block;
  margin: 0.55rem 0;
}}
.sb-kv .k {{
  color: rgba(0, 0, 0, 0.55);
  font-size: 14px;
  font-weight: 700;
  white-space: nowrap;
  margin-bottom: 0.25rem;
}}
.sb-kv .v {{
  font-size: 14px;
  line-height: 1.45;
  color: rgba(0, 0, 0, 0.85);
  overflow-wrap: anywhere;
}}
.sb-kv code {{
  white-space: normal;
}}
.sb-dot {{
  display: inline-block;
  width: 8px;
  height: 8px;
  border-radius: 999px;
  margin-right: 6px;
  background: #9aa0a6;
}}
.sb-dot.ok {{ background: #1e8e3e; }}
.sb-dot.warn {{ background: #b06000; }}

div[data-testid="stDownloadButton"] button {{
  width: 100%;
  border-radius: 10px !important;
  border: 1px solid rgba(0, 0, 0, 0.12) !important;
  background: rgba(255, 255, 255, 0.75) !important;
  font-weight: 800 !important;
  font-size: 14px !important;
  padding: 0.65rem 0.75rem !important;
}}
div[data-testid="stDownloadButton"] button:hover {{
  border-color: rgba(0, 0, 0, 0.20) !important;
}}
 /* Make title breathe */
 h1 {{
   margin-bottom: 0.8rem !important;
 }}

 /* DataFrame: match dashboard background (Streamlit DataTable / BaseWeb) */
 div[data-testid="stDataFrame"] {{
   background-color: {bg} !important;
 }}
 .stDataFrame {{
   background-color: {bg} !important;
 }}
 div[data-testid="stDataFrame"] div[data-baseweb="data-table"] {{
   background-color: {bg} !important;
 }}
 .stDataFrame div[data-baseweb="data-table"] {{
   background-color: {bg} !important;
 }}
 div[data-testid="stDataFrame"] div[data-baseweb="data-table"] div[role="gridcell"] {{
   background-color: {bg} !important;
 }}
 .stDataFrame div[data-baseweb="data-table"] div[role="gridcell"] {{
   background-color: {bg} !important;
 }}
 div[data-testid="stDataFrame"] div[data-baseweb="data-table"] div[role="row"] {{
   background-color: {bg} !important;
 }}
 .stDataFrame div[data-baseweb="data-table"] div[role="row"] {{
   background-color: {bg} !important;
 }}
 div[data-testid="stDataFrame"] div[data-baseweb="data-table"] div[role="columnheader"] {{
   background-color: {sbg} !important;
 }}
 .stDataFrame div[data-baseweb="data-table"] div[role="columnheader"] {{
   background-color: {sbg} !important;
 }}
 div[data-testid="stDataFrame"] div[data-baseweb="data-table"] div[role="row"]:hover div[role="gridcell"] {{
   background-color: {sbg} !important;
 }}
 .stDataFrame div[data-baseweb="data-table"] div[role="row"]:hover div[role="gridcell"] {{
   background-color: {sbg} !important;
 }}

 /* DataFrame: fallback for table-based renderers */
 .stDataFrame table {{
   background-color: {bg} !important;
 }}
 .stDataFrame thead tr th {{
   background-color: {sbg} !important;
 }}
 .stDataFrame tbody tr td {{
   background-color: {bg} !important;
 }}
 .stDataFrame tbody tr:hover td {{
   background-color: {sbg} !important;
 }}
 </style>
         """,
         unsafe_allow_html=True,
     )


def _split_family(family: str) -> tuple[str, str]:
    # Expected format: "<제품명> + <POWER>"
    if family is None:
        return "", ""
    s = str(family)
    if " + " not in s:
        return s, ""
    left, right = s.rsplit(" + ", 1)
    return left.strip(), right.strip()


def _format_int(x) -> str:
    try:
        v = float(x)
    except Exception:
        return "0"
    if pd.isna(v):
        return "0"
    return f"{int(round(v)):,}"


def _normalize_signed_2dp(v, *, zero_sign: str = "+") -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return f"{zero_sign}00.00"
    s = str(v).strip()
    if not s or s.lower() == "nan" or s == "<NA>":
        return f"{zero_sign}00.00"
    try:
        x = float(s)
    except Exception:
        return s
    if pd.isna(x):
        return f"{zero_sign}00.00"
    sign = zero_sign if x == 0 else ("-" if x < 0 else "+")
    mag = abs(x)
    return f"{sign}{mag:.2f}"


def _parse_search_terms(raw: str) -> list[str]:
    if raw is None:
        return []
    terms = [t.strip() for t in str(raw).split(",")]
    return [t for t in terms if t]


def _filter_by_name_contains(df: pd.DataFrame, name_col: str, raw_terms: str) -> pd.DataFrame:
    terms = _parse_search_terms(raw_terms)
    if not terms or name_col not in df.columns:
        return df
    # OR-match: include row if any term is contained (case-insensitive).
    pattern = "|".join(re.escape(t) for t in terms)
    mask = df[name_col].astype("string").fillna("").str.contains(pattern, case=False, regex=True, na=False)
    return df.loc[mask].copy()


def _filter_by_any_contains(df: pd.DataFrame, cols: list[str], raw_terms: str) -> pd.DataFrame:
    terms = _parse_search_terms(raw_terms)
    cols = [c for c in cols if c in df.columns]
    if not terms or not cols:
        return df
    pattern = "|".join(re.escape(t) for t in terms)
    mask = pd.Series(False, index=df.index)
    for c in cols:
        mask = mask | df[c].astype("string").fillna("").str.contains(pattern, case=False, regex=True, na=False)
    return df.loc[mask].copy()


DEFAULT_STAGE_COLS = ["사출", "분리", "하이드레이션", "접착", "누수규격"]


@st.cache_data(show_spinner=False)
def _to_excel_bytes(df: pd.DataFrame, *, sheet_name: str = "data") -> bytes:
    output = BytesIO()
    xdf = df.copy()

    # Ensure date columns export as YYYY-MM-DD (no time component).
    for c in xdf.columns:
        if c == "납기일":
            dt = pd.to_datetime(xdf[c], errors="coerce")
            xdf[c] = dt.dt.strftime("%Y-%m-%d")
        elif pd.api.types.is_datetime64_any_dtype(xdf[c]):
            xdf[c] = xdf[c].dt.strftime("%Y-%m-%d")
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            xdf.to_excel(writer, index=False, sheet_name=sheet_name[:31])
    except Exception:
        # Fallback to xlsxwriter if available, otherwise raise.
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:  # type: ignore[call-arg]
            xdf.to_excel(writer, index=False, sheet_name=sheet_name[:31])
    return output.getvalue()


def _to_injection_operation_xlsx(
    sched: pd.DataFrame,
    *,
    start_date: date,
    horizon_days: int,
    sheet_name: str = "운영양식",
    equip_all: pd.DataFrame | None = None,
) -> bytes:
    """
    Export injection schedule to an operator-friendly template-like Excel.

    Layout:
    - Column A: 설비 (merged across 8 cavity rows)
    - For each slot (day x block): 4 columns = 제품정보 / CAV / 도수 / 필요수량
    - Each equipment occupies 8 rows (CAV 1..8)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    thin = Side(style="thin", color="CFCFCF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="E8F0FE")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

    horizon_days = int(horizon_days)
    if horizon_days <= 0:
        horizon_days = 5

    slots: list[dict[str, object]] = []
    for i in range(horizon_days):
        d = start_date + timedelta(days=i)
        for b, sh in [(1, "주간"), (2, "야간")]:
            slots.append({"날짜": d, "Block": b, "shift": sh, "label": f"{d.month}/{d.day} {sh}"})

    # Header rows
    ws["A1"] = "설비"
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws["A1"].fill = header_fill
    ws["A1"].font = header_font
    ws["A1"].alignment = center

    col = 2
    for sl in slots:
        label = str(sl["label"])
        ws.cell(row=1, column=col, value=label)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
        for c in range(col, col + 4):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        sub = ["제품정보", "CAV", "도수", "필요수량"]
        for j, name in enumerate(sub):
            cell = ws.cell(row=2, column=col + j, value=name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        col += 4

    # Prepare schedule lookup
    s = sched.copy() if isinstance(sched, pd.DataFrame) else pd.DataFrame()
    if "날짜" in s.columns:
        s["날짜"] = pd.to_datetime(s["날짜"], errors="coerce").dt.date
    if "Block" in s.columns:
        s["Block"] = pd.to_numeric(s["Block"], errors="coerce").fillna(0).astype(int)
    if "설비명" in s.columns:
        s["설비명"] = s["설비명"].astype("string").fillna("").astype(str).str.strip().str.upper()

    key_cols = ["설비명", "날짜", "Block"]
    if not all(c in s.columns for c in key_cols):
        s_map: dict[tuple[str, date, int], dict[str, object]] = {}
    else:
        s_map = {
            (str(r["설비명"]), r["날짜"], int(r["Block"])): r.to_dict()
            for _, r in s.iterrows()
            if isinstance(r.get("날짜"), date)
        }

    def _equip_sort_key(s: str) -> tuple[int, int, str]:
        s2 = str(s or "").strip().upper()
        m = re.match(r"^([A-Z])(\d+)$", s2)
        if not m:
            return (999, 999, s2)
        return (ord(m.group(1)) - 65, int(m.group(2)), s2)

    equip_info: dict[str, dict[str, object]] = {}
    equip_list: list[str] = []
    if isinstance(equip_all, pd.DataFrame) and (not equip_all.empty):
        e = equip_all.copy()
        e["설비명"] = (
            e.get("설비코드", "")
            .astype("string")
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
        )
        e["비고"] = e.get("비고", "").astype("string").fillna("").astype(str).str.strip()
        e["생산 제품"] = e.get("생산 제품", "").astype("string").fillna("").astype(str).str.strip()
        e["현재제품코드"] = e.get("현재제품코드", "").astype("string").fillna("").astype(str).str.strip().str.upper()
        e["배정가능"] = e.get("배정가능", True).fillna(True)
        for _, r in e.iterrows():
            nm = str(r.get("설비명") or "").strip().upper()
            if not nm:
                continue
            equip_info[nm] = {
                "배정가능": bool(r.get("배정가능", True)),
                "비고": str(r.get("비고") or "").strip(),
                "현재제품": str(r.get("생산 제품") or "").strip(),
                "현재제품코드": str(r.get("현재제품코드") or "").strip().upper(),
            }
        equip_list = sorted(list(equip_info.keys()), key=_equip_sort_key)
    else:
        equip_list = sorted([e for e in s.get("설비명", pd.Series(dtype=str)).dropna().astype(str).unique().tolist() if str(e).strip()], key=_equip_sort_key)

    start_row = 3
    now = datetime.now(KST)
    now_block = 2 if now.hour >= 20 else 1
    for equip in equip_list:
        info = equip_info.get(str(equip), {})
        cur_code = str(info.get("현재제품코드") or "").strip().upper()
        cur_name = str(info.get("현재제품") or "").strip()
        # Fill running product across the whole horizon only when schedule has no assignments at all for this equipment.
        has_any_sched = False
        try:
            for sl in slots:
                d = sl["날짜"]
                b = int(sl["Block"])
                rec0 = s_map.get((str(equip), d, b), {})
                if str(rec0.get("제품명코드") or "").strip():
                    has_any_sched = True
                    break
        except Exception:
            has_any_sched = False
        fill_running_all = bool(cur_code) and (not has_any_sched)

        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + 7, end_column=1)
        c0 = ws.cell(row=start_row, column=1, value=str(equip))
        c0.alignment = center
        c0.font = Font(bold=True)

        # CAV rows
        for i in range(8):
            ws.cell(row=start_row + i, column=1).border = border

        col = 2
        for sl in slots:
            d = sl["날짜"]
            b = int(sl["Block"])
            rec = s_map.get((str(equip), d, b), {})

            prod = str(rec.get("제품명코드") or "").strip()
            prod_name = str(rec.get("제품명") or "").strip()

            # Keep running product shown only when:
            # - the whole horizon has no assignments for this equipment (fill_running_all), OR
            # - it's the current slot and schedule didn't assign anything.
            if (not prod) and fill_running_all and cur_code:
                prod = cur_code
                prod_name = cur_name
            elif (not prod) and isinstance(d, date) and (d == start_date) and (int(b) == int(now_block)) and cur_code:
                prod = cur_code
                prod_name = cur_name
            info_txt = "\n".join([t for t in [prod, prod_name] if t])
            if not info_txt:
                note = str(info.get("비고") or "").strip()
                assignable = bool(info.get("배정가능", True))
                if (not assignable) and note:
                    info_txt = f"배정불가\n비고: {note}"
                elif note:
                    info_txt = f"유휴\n비고: {note}"
                else:
                    info_txt = ""

            # 제품정보 (merged)
            ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row + 7, end_column=col)
            cinfo = ws.cell(row=start_row, column=col, value=info_txt)
            cinfo.alignment = left_top

            for i in range(8):
                r = start_row + i
                ws.cell(row=r, column=col + 1, value=i + 1).alignment = center
                pw = str(rec.get(f"PW{i + 1}") or "").strip()
                qv = rec.get(f"Q{i + 1}")
                try:
                    q = int(qv) if qv is not None and str(qv).strip() != "" else 0
                except Exception:
                    q = 0
                ws.cell(row=r, column=col + 2, value=pw).alignment = center
                ws.cell(row=r, column=col + 3, value=(q if q > 0 else "")).alignment = center

            # Borders for group
            for r in range(start_row, start_row + 8):
                for c in range(col, col + 4):
                    ws.cell(row=r, column=c).border = border
            col += 4

        start_row += 8

    # Column widths
    ws.column_dimensions["A"].width = 8
    col = 2
    for _ in slots:
        ws.column_dimensions[get_column_letter(col)].width = 22  # 제품정보
        ws.column_dimensions[get_column_letter(col + 1)].width = 5  # CAV
        ws.column_dimensions[get_column_letter(col + 2)].width = 8  # WET
        ws.column_dimensions[get_column_letter(col + 3)].width = 9  # 수량
        col += 4

    # Apply borders to headers
    max_col = 1 + 4 * len(slots)
    for r in (1, 2):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = border

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


@st.cache_data(show_spinner=False)
def _to_injection_operation_xlsx_cached(
    sched: pd.DataFrame,
    *,
    start_date: date,
    horizon_days: int,
    sheet_name: str,
    equip_all: pd.DataFrame | None,
    excel_mtime: float,
    now_block: int,
) -> bytes:
    # Cache-buster: regenerate when the raw APS Excel changes.
    _ = float(excel_mtime)
    _ = int(now_block)
    return _to_injection_operation_xlsx(
        sched,
        start_date=start_date,
        horizon_days=horizon_days,
        sheet_name=sheet_name,
        equip_all=equip_all,
    )


def _injection_schedule_to_cavity_rows(sched: pd.DataFrame) -> pd.DataFrame:
    """
    Expand block schedule rows into cavity-level rows.

    Output columns:
    - 날짜, 주야, 설비명, Block, 제품명코드, 제품명, CAV, 도수, 필요수량
    """
    if sched is None or sched.empty:
        return pd.DataFrame()

    df = sched.copy()
    if "날짜" not in df.columns or "Block" not in df.columns or "설비명" not in df.columns:
        return pd.DataFrame()

    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce").dt.date
    df["Block"] = pd.to_numeric(df["Block"], errors="coerce").fillna(0).astype(int)
    df["주야"] = df["Block"].map(lambda b: "주간" if int(b) == 1 else ("야간" if int(b) == 2 else ""))

    rows: list[dict[str, object]] = []
    for _, r in df.iterrows():
        d = r.get("날짜")
        if not isinstance(d, date):
            continue
        equip = str(r.get("설비명") or "").strip().upper()
        if not equip:
            continue
        block = int(r.get("Block") or 0)
        sh = str(r.get("주야") or "").strip()
        prod = str(r.get("제품명코드") or "").strip().upper()
        prod_name = str(r.get("제품명") or "").strip()
        for i in range(8):
            cav = i + 1
            pw = str(r.get(f"PW{cav}") or "").strip()
            qv = r.get(f"Q{cav}")
            try:
                q = int(qv) if qv is not None and str(qv).strip() != "" else 0
            except Exception:
                q = 0
            if (not pw) and q <= 0:
                continue
            rows.append(
                {
                    "날짜": d,
                    "주야": sh,
                    "설비명": equip,
                    "Block": block,
                    "제품명코드": prod,
                    "제품명": prod_name,
                    "CAV": cav,
                    "도수": pw,
                    "필요수량": q,
                }
            )

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["필요수량"] = pd.to_numeric(out["필요수량"], errors="coerce").fillna(0).astype(int)
    out = out.sort_values(["설비명", "날짜", "Block", "CAV"], ascending=[True, True, True, True], na_position="last")
    return out


def _build_injection_capacity_segments(
    sched: pd.DataFrame,
    *,
    start_date: date,
    horizon_days: int,
) -> list[dict[str, object]]:
    """
    Build a piecewise capacity timeline from injection schedule.

    Time unit: "days since start_date".
    Each block is 0.5 day: Block1 -> [d+0.0, d+0.5), Block2 -> [d+0.5, d+1.0).
    Segment capacity is total assigned qty within that block across all equipments.
    """
    if sched is None or sched.empty:
        return []

    df = sched.copy()
    if "날짜" not in df.columns or "Block" not in df.columns:
        return []

    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce").dt.date
    df["Block"] = pd.to_numeric(df["Block"], errors="coerce").fillna(0).astype(int)
    df["배정수량"] = pd.to_numeric(df.get("배정수량", 0), errors="coerce").fillna(0).astype(int)

    horizon_days = int(horizon_days)
    if horizon_days <= 0:
        horizon_days = 5

    segs: list[dict[str, object]] = []
    for i in range(horizon_days):
        d = start_date + timedelta(days=i)
        for b, start_off in [(1, 0.0), (2, 0.5)]:
            cap = int(df.loc[(df["날짜"] == d) & (df["Block"] == int(b)), "배정수량"].sum())
            segs.append(
                {
                    "t0": float(i) + float(start_off),
                    "dur": 0.5,
                    "cap": int(max(0, cap)),
                }
            )
    return segs


@st.cache_data(show_spinner=False)
def _build_injection_plan_segments_cached(
    *,
    demand: pd.DataFrame,
    inj_equip: pd.DataFrame,
    inj_arrange: pd.DataFrame,
    excel_path: str,
    excel_mtime: float,
    start_date: date,
    horizon_days: int,
) -> list[dict[str, object]]:
    inj_sched, _, _ = _build_injection_schedule_cached(
        demand=demand,
        inj_equip=inj_equip,
        arrange=inj_arrange,
        excel_path=excel_path,
        excel_mtime=excel_mtime,
        start_date=start_date,
        horizon_days=horizon_days,
    )
    if inj_sched is None or inj_sched.empty:
        return []
    return _build_injection_capacity_segments(inj_sched, start_date=start_date, horizon_days=horizon_days)


class _CapacityAllocator:
    def __init__(self, segs: list[dict[str, object]]):
        self.segs = segs or []
        self.i = 0
        self.t = float(self.segs[0]["t0"]) if self.segs else 0.0
        self.rem = int(self.segs[0]["cap"]) if self.segs else 0

    def _advance_to(self, t: float) -> None:
        # Move to the segment that contains or follows time t.
        if not self.segs:
            self.t = float(t)
            self.i = 0
            self.rem = 0
            return
        while self.i < len(self.segs):
            seg = self.segs[self.i]
            t0 = float(seg["t0"])
            dur = float(seg["dur"])
            t1 = t0 + dur
            if float(t) < t0:
                self.t = t0
                self.rem = int(seg.get("cap") or 0)
                return
            if t0 <= float(t) < t1:
                self.t = float(t)
                # Remaining capacity in this segment: proportional to remaining time.
                cap = int(seg.get("cap") or 0)
                frac = max(0.0, (t1 - float(t)) / dur) if dur > 0 else 0.0
                self.rem = int(round(cap * frac))
                return
            self.i += 1
        # Beyond last segment
        last = self.segs[-1]
        self.t = float(last["t0"]) + float(last["dur"])
        self.rem = 0

    def allocate(self, qty: float, *, earliest_start: float) -> float:
        """
        Allocate `qty` units onto the capacity timeline.
        Returns done time (days since start_date).
        """
        q = float(qty or 0.0)
        if q <= 0:
            return float(max(self.t, float(earliest_start)))

        start = float(max(self.t, float(earliest_start)))
        self._advance_to(start)
        if not self.segs:
            # No capacity info; treat as no-op duration.
            self.t = start
            return start

        while q > 0 and self.i < len(self.segs):
            seg = self.segs[self.i]
            t0 = float(seg["t0"])
            dur = float(seg["dur"])
            t1 = t0 + dur
            if self.t < t0:
                self.t = t0
                self.rem = int(seg.get("cap") or 0)

            if self.rem <= 0:
                self.i += 1
                if self.i < len(self.segs):
                    self.t = float(self.segs[self.i]["t0"])
                    self.rem = int(self.segs[self.i].get("cap") or 0)
                continue

            take = min(float(self.rem), q)
            q -= take
            self.rem -= int(round(take))

            if q <= 0:
                return float(self.t)

            # Move to next segment boundary when current segment is exhausted.
            if self.rem <= 0:
                self.t = t1
                self.i += 1
                if self.i < len(self.segs):
                    self.t = float(self.segs[self.i]["t0"])
                    self.rem = int(self.segs[self.i].get("cap") or 0)

        # If we run out of segments, finish at end of last segment.
        if self.segs:
            last = self.segs[-1]
            self.t = float(last["t0"]) + float(last["dur"])
        return float(self.t)


def _order_ref_string(ini: object, order_no: object) -> str:
    i = str(ini or "").strip()
    o = str(order_no or "").strip()
    # Convention (수주별 현황): if initials exist, identify by initials; otherwise use order number
    # (e.g., 해외/국내/PB 등은 이니셜이 비어 수주번호로 구분되는 경우가 많음).
    return i or o


def _format_order_ref_list(items: list[tuple[date, str]], *, max_show: int = 10) -> str:
    if not items:
        return ""
    items2 = [(d if isinstance(d, date) else date(2099, 12, 31), str(k or "").strip()) for d, k in items]
    items2 = [(d, k) for d, k in items2 if k]
    items2.sort(key=lambda t: (t[0], t[1]))
    keys = [k for _, k in items2]
    if not keys:
        return ""
    if len(keys) <= int(max_show):
        return ", ".join(keys)
    return ", ".join(keys[: int(max_show)]) + f" …(+{len(keys) - int(max_show)})"


@st.cache_data(show_spinner=False)
def _build_order_refs_by_base_r(detail: pd.DataFrame) -> dict[str, list[tuple[date, str]]]:
    """
    Build base R -> list of (min_due, order_ref) from order-detail.
    order_ref = "이니셜-수주번호" (fallback to whichever exists).
    """
    if detail is None or detail.empty or ("제품 코드" not in detail.columns):
        return {}

    d = detail.copy()
    d["제품 코드"] = d["제품 코드"].astype("string").fillna("").astype(str).str.strip().str.upper()
    d = d.loc[d["제품 코드"].str.startswith("R")].copy()
    if d.empty:
        return {}

    d["base_r"] = d["제품 코드"].map(_extract_base_r).astype("string").fillna("").astype(str).str.strip().str.upper()
    d = d.loc[d["base_r"].ne("")].copy()
    if d.empty:
        return {}

    if "납기일" in d.columns:
        d["납기일"] = pd.to_datetime(d["납기일"], errors="coerce").dt.date
    else:
        d["납기일"] = None

    if "이니셜" not in d.columns:
        d["이니셜"] = ""
    if "수주번호" not in d.columns:
        d["수주번호"] = ""

    d["order_ref"] = d.apply(lambda r: _order_ref_string(r.get("이니셜"), r.get("수주번호")), axis=1)
    d = d.loc[d["order_ref"].astype("string").fillna("").astype(str).str.strip().ne("")].copy()
    if d.empty:
        return {}

    # base_r + order_ref -> min due
    agg = (
        d.groupby(["base_r", "order_ref"], dropna=False, as_index=False)["납기일"]
        .min()
        .rename(columns={"납기일": "min_due"})
    )
    out: dict[str, list[tuple[date, str]]] = {}
    for _, r in agg.iterrows():
        br = str(r.get("base_r") or "").strip().upper()
        ref = str(r.get("order_ref") or "").strip()
        due = r.get("min_due")
        if not br or not ref:
            continue
        out.setdefault(br, []).append((_power_due_or_far(due), ref))
    return out


def _format_item_code_list(codes: list[str], *, max_show: int = 12) -> str:
    codes = [str(c).strip() for c in codes if str(c).strip()]
    if not codes:
        return ""
    uniq = []
    seen = set()
    for c in codes:
        if c in seen:
            continue
        uniq.append(c)
        seen.add(c)
    if len(uniq) <= max_show:
        return ", ".join(uniq)
    head = ", ".join(uniq[:max_show])
    return f"{head} …(+{len(uniq) - max_show})"


def _attach_item_codes(
    df: pd.DataFrame,
    detail: pd.DataFrame | None,
    *,
    out_col: str = "제품코드",
    allowed_prefixes: list[str] | None = None,
) -> pd.DataFrame:
    if detail is None or detail.empty:
        return df
    if "제품 코드" not in detail.columns:
        return df

    # Join keys:
    # Use lens-spec columns (ADD/CP/AXIS) so Toric/Multi rows map to the correct item code.
    # When a column is missing on either side, create an empty normalized column to avoid
    # accidental broad joins (which can cause mixed/duplicated codes in one cell).
    key_cols = ["신규분류 요약코드", "제품군", "ADD", "CP", "AXIS", "납기일"]
    required = [c for c in key_cols if c != "납기일"]

    d = detail.copy()
    if "납기일" in d.columns:
        d["납기일"] = pd.to_datetime(d["납기일"], errors="coerce")
    left = df.copy()
    if "납기일" in left.columns:
        left["납기일"] = pd.to_datetime(left["납기일"], errors="coerce")

    # Ensure required string keys exist on both sides.
    for c in required:
        if c not in d.columns:
            d[c] = ""
        if c not in left.columns:
            left[c] = ""

    # Normalize string keys to avoid mismatch due to NA vs "" vs whitespace.
    for c in key_cols:
        if c == "납기일":
            continue
        d[c] = d[c].astype("string").fillna("").str.strip()
        left[c] = left[c].astype("string").fillna("").str.strip()

    prefixes: list[str] | None = None
    if allowed_prefixes:
        prefixes = [str(p).strip().upper() for p in allowed_prefixes if str(p).strip()]

    def _codes_for_group(s: pd.Series) -> str:
        items = s.dropna().astype(str).map(lambda x: x.strip())
        if prefixes:
            items = items[items.map(lambda x: any(x.upper().startswith(p) for p in prefixes))]
        # If there are still multiple codes, try to match the group AXIS (last 3 digits in code).
        try:
            group_key = s.name if isinstance(s.name, tuple) else (s.name,)
            axis_idx = key_cols.index("AXIS")
            axis_val = str(group_key[axis_idx]).strip()
        except Exception:
            axis_val = ""
        if axis_val.isdigit() and len(axis_val) == 3 and len(items) > 1:
            items_axis = items[items.map(lambda x: str(x)[-3:].isdigit() and str(x)[-3:] == axis_val)]
            if len(items_axis) > 0:
                items = items_axis

        uniq = sorted(set([x for x in items.tolist() if str(x).strip()]))
        if not uniq:
            return ""
        # 공정별 보기에서는 한 행=한 제품코드가 되도록 1개로 고정 표시
        return uniq[0]

    codes = (
        d.groupby(key_cols, dropna=False)["제품 코드"]
        .apply(_codes_for_group)
        .reset_index()
        .rename(columns={"제품 코드": out_col})
    )
    merged = left.merge(codes, on=key_cols, how="left")
    return merged


@st.cache_data(show_spinner=False)
def _load_due_csv(path: str, mtime: float) -> pd.DataFrame:
    _ = mtime  # cache-buster when file changes
    header = pd.read_csv(path, nrows=0)
    dtype: dict[str, str] = {}
    for c in ["신규분류 요약코드", "제품군", "ADD", "CP", "AXIS"]:
        if c in header.columns:
            dtype[c] = "string"
    df = pd.read_csv(path, dtype=dtype if dtype else None)
    df["납기일"] = pd.to_datetime(df["납기일"], errors="coerce")
    return df


@st.cache_data(show_spinner=False)
def _load_order_detail_csv(path: str, mtime: float) -> pd.DataFrame:
    _ = mtime  # cache-buster when file changes
    header = pd.read_csv(path, nrows=0)
    dtype: dict[str, str] = {}
    for c in ["이니셜", "수주번호", "신규분류 요약코드", "수요 제품 이름", "제품군", "제품 코드", "ADD", "CP", "AXIS"]:
        if c in header.columns:
            dtype[c] = "string"
    df = pd.read_csv(path, dtype=dtype if dtype else None)
    if "납기일" in df.columns:
        df["납기일"] = pd.to_datetime(df["납기일"], errors="coerce")
    return df


@st.cache_data(show_spinner=False)
def _load_equip_code_min_target_csv(path: str, mtime: float) -> pd.DataFrame:
    _ = mtime  # cache-buster when file changes
    if not path or not os.path.exists(path):
        return pd.DataFrame()
    header = pd.read_csv(path, nrows=0)
    dtype: dict[str, str] = {}
    for c in ["공정", "제품 코드"]:
        if c in header.columns:
            dtype[c] = "string"
    df = pd.read_csv(path, dtype=dtype if dtype else None)
    if "최소목표일" in df.columns:
        df["최소목표일"] = pd.to_datetime(df["최소목표일"], errors="coerce")
    for c in ["공정", "제품 코드"]:
        if c in df.columns:
            df[c] = df[c].astype("string").fillna("").str.strip()
    return df


@st.cache_data(show_spinner=False)
def _load_prod_daily_csv(path: str | None, mtime: float) -> pd.DataFrame:
    _ = mtime  # cache-buster when file changes
    if not path or not os.path.exists(path):
        return pd.DataFrame()
    header = pd.read_csv(path, nrows=0)
    dtype: dict[str, str] = {}
    for c in ["공정"]:
        if c in header.columns:
            dtype[c] = "string"
    df = pd.read_csv(path, dtype=dtype if dtype else None)
    if "생산일자" in df.columns:
        df["생산일자"] = pd.to_datetime(df["생산일자"], errors="coerce")
    for c in ["공정"]:
        if c in df.columns:
            df[c] = df[c].astype("string").fillna("").str.strip()
    if "양품" in df.columns:
        df["양품"] = pd.to_numeric(df["양품"], errors="coerce").fillna(0)
    return df


@st.cache_data(show_spinner=False)
def _load_injection_sheet_cached(path: str, mtime: float) -> dict[str, pd.DataFrame]:
    _ = mtime  # cache-buster when file changes
    try:
        raw = pd.read_excel(path, sheet_name="사출")
    except Exception:
        return {"equip": pd.DataFrame(), "arrange": pd.DataFrame()}

    # Note: in current xlsx, columns map as:
    # - 사출 호기: A1/A2... (equipment code)
    # - 구분: 라인구분(예: "S관 사출조립 인라인")
    # - 구분2: 설비 타입(예: CLEAR/COLOR)
    equip_cols = ["위치", "사출 호기", "구분", "구분2", "생산 제품", "비고"]
    if "사출 호기" in raw.columns:
        equip = raw.loc[raw["사출 호기"].notna(), [c for c in equip_cols if c in raw.columns]].copy()
    else:
        equip = pd.DataFrame(columns=[c for c in equip_cols if c in raw.columns])
    if not equip.empty and "사출 호기" in equip.columns:
        equip["설비코드"] = (
            equip["사출 호기"]
            .astype("string")
            .fillna("")
            .astype(str)
            .str.strip()
            .str.split()
            .str[0]
        )
    else:
        equip["설비코드"] = ""

    def _is_blank(v: object) -> bool:
        if v is None:
            return True
        s = str(v).strip()
        return (s == "") or (s.lower() == "nan")

    if "생산 제품" not in equip.columns:
        equip["생산 제품"] = ""
    if "비고" not in equip.columns:
        equip["비고"] = ""
    equip["생산 제품"] = equip["생산 제품"].astype("string").fillna("").astype(str).str.strip()
    equip["비고"] = equip["비고"].astype("string").fillna("").astype(str).str.strip()

    # Rule: E(생산 제품) 공란 + F(비고) 기입 => 배정 불가
    equip["배정가능"] = ~equip.apply(lambda r: _is_blank(r.get("생산 제품")) and (not _is_blank(r.get("비고"))), axis=1)

    arrange_cols = ["제품명코드", "제품명", "구분.1"]
    if "제품명코드" in raw.columns:
        arrange = raw.loc[raw["제품명코드"].notna(), [c for c in arrange_cols if c in raw.columns]].copy()
    else:
        arrange = pd.DataFrame(columns=[c for c in arrange_cols if c in raw.columns])
    if arrange.empty:
        arrange = pd.DataFrame(columns=["제품명코드", "제품명", "구분.1"])
    for c in ["제품명코드", "제품명", "구분.1"]:
        if c not in arrange.columns:
            arrange[c] = ""
        arrange[c] = arrange[c].astype("string").fillna("").astype(str).str.strip()

    def _to_base_r(v: object) -> str:
        s = str(v or "").strip().upper()
        m = re.match(r"^(R\d+)", s)
        return m.group(1) if m else s

    # Allow users to paste full item code (e.g., R1026+03.75...) but normalize to base R (R1026)
    arrange["제품명코드"] = arrange["제품명코드"].map(_to_base_r)

    # Map current running product name -> base R code for display (간트/상세표).
    try:
        if (not equip.empty) and (not arrange.empty) and ("생산 제품" in equip.columns) and ("제품명" in arrange.columns):
            name_to_base: dict[str, str] = {}
            for _, r in arrange.iterrows():
                nm = _norm_space(r.get("제품명"))
                br = str(r.get("제품명코드") or "").strip().upper()
                if nm and br and (nm not in name_to_base):
                    name_to_base[nm] = br
            equip["현재제품코드"] = equip["생산 제품"].map(lambda v: name_to_base.get(_norm_space(v), ""))
        else:
            equip["현재제품코드"] = ""
    except Exception:
        equip["현재제품코드"] = ""
    return {"equip": equip, "arrange": arrange}


@st.cache_data(show_spinner=False)
def _load_injection_machine_medians_cached(path: str, mtime: float) -> dict[str, object]:
    _ = mtime  # cache-buster when file changes
    try:
        df = pd.read_excel(path, sheet_name="생산실적", usecols=["생산일자", "공정코드", "양품수량", "기계코드"])
    except Exception:
        return {"med_by_machine": {}, "global_median": 0.0}

    df["공정코드"] = df["공정코드"].astype("string").fillna("").astype(str)
    df = df.loc[df["공정코드"].str.contains("사출", na=False)].copy()
    if df.empty:
        return {"med_by_machine": {}, "global_median": 0.0}

    df["생산일자"] = pd.to_datetime(df["생산일자"], errors="coerce").dt.date
    df["양품수량"] = pd.to_numeric(df["양품수량"], errors="coerce").fillna(0.0)
    df["기계코드"] = df["기계코드"].astype("string").fillna("").astype(str).str.strip()
    df = df.loc[df["생산일자"].notna() & df["기계코드"].ne("")].copy()
    if df.empty:
        return {"med_by_machine": {}, "global_median": 0.0}

    daily = df.groupby(["기계코드", "생산일자"], dropna=False, as_index=False)["양품수량"].sum()
    med_by_machine = daily.groupby("기계코드", dropna=False)["양품수량"].median().to_dict()
    global_median = float(daily["양품수량"].median()) if not daily.empty else 0.0
    return {"med_by_machine": med_by_machine, "global_median": global_median}


def _infer_machine_code_from_equip(equip_code: str) -> str | None:
    s = str(equip_code or "").strip().upper()
    m = re.match(r"^([A-Z])(\d+)$", s)
    if not m:
        return None
    letter, n = m.group(1), int(m.group(2))
    if letter == "A":
        return f"A형 인라인{n}호기 - 조립중합"
    if letter == "B":
        return f"B형 조립중합{n}호기"
    if letter == "C":
        return f"C형 조립중합{n}호기"
    return None


def _extract_base_r(code: object) -> str:
    s = str(code or "").strip()
    if not s:
        return ""
    m = re.match(r"^(R\d+)", s, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip().upper()
    return s.split("-", 1)[0].strip().upper()


def _coerce_date_value(x: object) -> date | None:
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    t = pd.to_datetime(x, errors="coerce")
    if pd.isna(t):
        return None
    try:
        return t.date()
    except Exception:
        return None


def _choose_power_slots(powers_rem: dict[float, int], *, slots: int, slot_qty: int) -> tuple[list[float], list[int]]:
    """
    Return (slot_powers, slot_qtys) where each slot is assigned to one POWER (duplicates allowed).
    slot_qtys are capped by remaining qty; unused capacity is 0.
    """
    out_p: list[float] = []
    out_q: list[int] = []
    if slots <= 0 or slot_qty <= 0 or not powers_rem:
        return out_p, out_q

    def _best_candidates() -> list[tuple[float, int]]:
        items = [(p, int(q)) for p, q in powers_rem.items() if int(q) > 0]
        items.sort(key=lambda t: (-t[1], t[0]))
        return items

    for _ in range(int(slots)):
        cand = _best_candidates()
        if not cand:
            break
        p, rem = cand[0]
        q = int(min(rem, slot_qty))
        out_p.append(float(p))
        out_q.append(int(q))
        powers_rem[p] = int(rem - q)

    return out_p, out_q


def _fmt_power(p: float | None) -> str:
    if p is None:
        return ""
    try:
        return f"{float(p):+.2f}"
    except Exception:
        return ""


def _power_due_or_far(d: object) -> date:
    return d if isinstance(d, date) else date(2099, 12, 31)


def _pick_power_types_for_block(
    powers: dict[float, dict[str, object]],
    *,
    prev_types: list[float],
    block_day: date,
    max_types: int,
) -> list[float]:
    """
    Decide which POWER types to use for the current block.

    Goals:
    - Prefer continuing previous block's POWER if remaining.
    - Reflect POWER-level due date (earlier first).
    - Minimize POWER variety within a block (default max 2).
    - Allow exceptions when there are urgent POWERs (due <= block_day).
    """
    avail: list[tuple[float, date, int]] = []
    for p, info in (powers or {}).items():
        try:
            qty = int(info.get("qty") or 0)
        except Exception:
            qty = 0
        if qty <= 0:
            continue
        due = _power_due_or_far(info.get("due"))
        avail.append((float(p), due, qty))

    if not avail:
        return []

    # Best-first by due, then remaining qty, then POWER numeric.
    avail_sorted = sorted(avail, key=lambda t: (t[1], -t[2], t[0]))
    urgent = [p for (p, due, qty) in avail_sorted if (qty > 0 and due <= block_day)]

    prev_val: float | None = None
    if prev_types:
        try:
            prev_val = float(prev_types[0])
        except Exception:
            prev_val = None

    chosen: list[float] = []

    # 1) Primary POWER: due date is the top priority.
    # Always select from the earliest-due group; only keep previous POWER when it belongs to that same group.
    min_due = avail_sorted[0][1]
    earliest = [(p, due, qty) for (p, due, qty) in avail_sorted if due == min_due]

    prev_due: date | None = None
    if isinstance(prev_val, float):
        for p, d, _ in avail_sorted:
            if float(p) == float(prev_val):
                prev_due = d
                break

    if isinstance(prev_val, float) and (prev_due == min_due) and any(float(prev_val) == float(p) for (p, _, _) in earliest):
        chosen.append(float(prev_val))
    else:
        # Prefer near previous POWER to reduce oscillation, but only within the earliest-due group.
        cand = earliest
        if isinstance(prev_val, float):
            cand.sort(key=lambda t: (abs(float(t[0]) - float(prev_val)), -t[2], t[0]))
        else:
            cand.sort(key=lambda t: (-t[2], t[0]))
        chosen.append(float(cand[0][0]))

    # 2) Secondary POWER (optional): prefer earliest-due group first, then proximity to primary.
    if int(max_types) >= 2:
        rem = [(p, due, qty) for (p, due, qty) in avail_sorted if float(p) not in set(chosen)]
        if rem:
            primary = float(chosen[0])
            rem_earliest = [t for t in rem if t[1] == min_due]
            cand2 = rem_earliest if rem_earliest else rem
            cand2.sort(key=lambda t: (t[1], abs(float(t[0]) - primary), -t[2], t[0]))
            chosen.append(float(cand2[0][0]))

    # 3) Exception: ensure urgent POWERs are included even if it increases variety.
    for p in urgent:
        if float(p) not in chosen:
            chosen.append(float(p))

    return chosen


def _choose_power_slots_min_change(
    powers: dict[float, dict[str, object]],
    *,
    prev_types: list[float],
    block_day: date,
    slots: int,
    slot_qty: int,
    max_types: int = 2,
    min_new_type_qty: int | None = None,
) -> tuple[list[float], list[int], list[float]]:
    """
    Allocate up to `slots` slots within a block while minimizing POWER changes.

    Returns (slot_powers, slot_qtys, chosen_types).
    - Slots are filled contiguously by chosen POWER types (no alternating), to reduce core changes.
    - Remaining capacity is left unused (0) rather than introducing extra POWER types, unless urgent.
    """
    out_p: list[float] = []
    out_q: list[int] = []
    if slots <= 0 or slot_qty <= 0 or not powers:
        return out_p, out_q, []

    chosen_types = _pick_power_types_for_block(powers, prev_types=prev_types, block_day=block_day, max_types=max(1, int(max_types)))
    if not chosen_types:
        return out_p, out_q, []

    prev_set = {float(p) for p in (prev_types or []) if isinstance(p, (int, float))}
    prev_ref: float | None = None
    if prev_types:
        try:
            prev_ref = float(prev_types[0])
        except Exception:
            prev_ref = None

    def _alloc_key(p: float) -> tuple[int, date, int, int, float]:
        info = powers.get(float(p)) or {}
        due = _power_due_or_far(info.get("due"))
        try:
            qty = int(info.get("qty") or 0)
        except Exception:
            qty = 0
        urgent_rank = 0 if due <= block_day else 1
        prev_rank = 0 if float(p) in prev_set else 1
        dist = 0
        if isinstance(prev_ref, float):
            dist = int(round(abs(float(p) - float(prev_ref)) * 1000))
        return (urgent_rank, due, prev_rank, dist, -qty, float(p))

    alloc_order = sorted([float(p) for p in chosen_types], key=_alloc_key)

    # Avoid "micro-run" POWERs that cause extra core changes but produce too little.
    # Policy: do not introduce a new POWER type in a block unless it can produce at least `min_new_type_qty`
    # within that block, unless it's urgent (due <= block_day).
    min_new = int(min_new_type_qty) if (min_new_type_qty is not None) else int(slot_qty)
    if min_new < 0:
        min_new = 0

    for p in alloc_order:
        if len(out_p) >= int(slots):
            break
        info = powers.get(float(p)) or {}
        due = _power_due_or_far(info.get("due"))
        try:
            qty0 = int(info.get("qty") or 0)
        except Exception:
            qty0 = 0

        is_urgent = bool(isinstance(due, date) and due <= block_day)
        is_prev = float(p) in prev_set
        is_primary = bool(out_p == [])
        introducing_new = (not is_prev) and (not is_primary)
        if introducing_new and (not is_urgent) and (qty0 < min_new):
            # Skip small remainder and keep capacity unused instead of adding another POWER.
            continue
        while len(out_p) < int(slots):
            try:
                rem = int(info.get("qty") or 0)
            except Exception:
                rem = 0
            if rem <= 0:
                break
            q = int(min(rem, int(slot_qty)))
            out_p.append(float(p))
            out_q.append(int(q))
            info["qty"] = int(rem - q)
        powers[float(p)] = info

    return out_p, out_q, chosen_types


def _choose_power_slots_for_8cav(
    powers: dict[float, dict[str, object]],
    *,
    prev_powers: list[float],
    block_day: date,
    slots: int,
    slot_qty: int,
    max_types: int = 8,
) -> tuple[list[float], list[int], list[float]]:
    """
    Choose per-cavity POWERs for an 8-cavity mold within a block.

    Key behavior (operator-friendly):
    - Allow up to 8 POWER types in a block (one per cavity).
    - Prefer continuing previous block's POWER set to reduce re-setting.
    - Prioritize urgent/early-due POWERs, then larger remaining qty.
    - Each cavity produces up to `slot_qty` (e.g. 2,000ea) and may produce less if remaining is smaller.
    """
    out_p: list[float] = []
    out_q: list[int] = []
    if slots <= 0 or slot_qty <= 0 or not powers:
        return out_p, out_q, []

    max_types = max(1, min(int(max_types), int(slots)))

    def _qty(p: float) -> int:
        try:
            return int((powers.get(float(p)) or {}).get("qty") or 0)
        except Exception:
            return 0

    def _due(p: float) -> date:
        info = powers.get(float(p)) or {}
        return _power_due_or_far(info.get("due"))

    avail = [float(p) for p in powers.keys() if _qty(float(p)) > 0]
    if not avail:
        return out_p, out_q, []

    prev_list = [float(p) for p in (prev_powers or []) if isinstance(p, (int, float))]
    prev_set = {float(p) for p in prev_list}

    # 1) Start with previous powers (keep order), if still needed.
    chosen: list[float] = []
    for p in prev_list:
        if len(chosen) >= max_types:
            break
        if float(p) in chosen:
            continue
        if _qty(float(p)) > 0:
            chosen.append(float(p))

    # 2) Add urgent powers (due <= block_day), then fill by due/qty.
    urgent = [p for p in avail if _due(float(p)) <= block_day]
    urgent.sort(key=lambda p: (_due(float(p)), -_qty(float(p)), float(p)))
    for p in urgent:
        if len(chosen) >= max_types:
            break
        if float(p) not in chosen:
            chosen.append(float(p))

    # 3) Fill remaining slots by due then remaining qty (and mild proximity to first prev power if present).
    ref = prev_list[0] if prev_list else None

    def _rank(p: float) -> tuple[date, int, int, float]:
        d = _due(float(p))
        q = _qty(float(p))
        dist = 0
        if isinstance(ref, float):
            dist = int(round(abs(float(p) - float(ref)) * 1000))
        prev_rank = 0 if float(p) in prev_set else 1
        return (d, prev_rank, dist, -q, float(p))

    rest = [p for p in avail if float(p) not in chosen]
    rest.sort(key=_rank)
    for p in rest:
        if len(chosen) >= max_types:
            break
        chosen.append(float(p))

    # Allocate 1 cavity per chosen type first (to avoid splitting across shifts).
    for p in chosen:
        if len(out_p) >= int(slots):
            break
        rem = _qty(float(p))
        if rem <= 0:
            continue
        q = int(min(rem, int(slot_qty)))
        out_p.append(float(p))
        out_q.append(int(q))
        info = powers.get(float(p)) or {}
        info["qty"] = int(rem - q)
        powers[float(p)] = info

    # Use remaining cavities for the most-remaining chosen powers (duplicates) when meaningful.
    while len(out_p) < int(slots):
        # Pick a power with the largest remaining among chosen.
        cand = sorted([p for p in chosen if _qty(float(p)) > 0], key=lambda p: (-_qty(float(p)), _due(float(p)), float(p)))
        if not cand:
            break
        p = float(cand[0])
        rem = _qty(p)
        q = int(min(rem, int(slot_qty)))
        out_p.append(p)
        out_q.append(q)
        info = powers.get(p) or {}
        info["qty"] = int(rem - q)
        powers[p] = info

    return out_p, out_q, chosen


def _build_injection_schedule(
    *,
    demand: pd.DataFrame,
    inj_equip: pd.DataFrame,
    arrange: pd.DataFrame,
    excel_path: str,
    excel_mtime: float,
    start_date: date,
    horizon_days: int,
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    """
    Build 4~5 day injection schedule (2 blocks/day/equipment, 8 power slots/block).
    Returns (schedule_df, remaining_df, warnings).
    """
    warnings: list[str] = []
    if demand is None or demand.empty:
        return (
            pd.DataFrame(
                columns=[
                    "날짜",
                    "설비명",
                    "제품명코드",
                    "제품명",
                    "Block",
                    "POWER 리스트",
                    "POWER 개수",
                    "배정수량",
                    "잔여수량",
                    "세팅구분",
                ]
            ),
            pd.DataFrame(columns=["제품명코드", "제품명", "납기일", "잔여수량"]),
            ["사출 부족수량(수요)이 없습니다."],
        )

    inj_equip = inj_equip.copy()
    arrange = arrange.copy()

    if "설비코드" not in inj_equip.columns:
        warnings.append("사출 시트 설비 테이블에서 설비코드를 찾지 못했습니다.")
        return (pd.DataFrame(), pd.DataFrame(), warnings)

    def _norm_space(v: object) -> str:
        s = str(v or "").strip()
        if not s:
            return ""
        return " ".join(s.split())

    arrange_map: dict[str, str] = {}
    arrange_name_map: dict[str, str] = {}
    if not arrange.empty and ("제품명코드" in arrange.columns):
        for _, r in arrange.iterrows():
            k = str(r.get("제품명코드") or "").strip().upper()
            if not k:
                continue
            v = _norm_space(r.get("구분.1")) if "구분.1" in arrange.columns else ""
            nm = _norm_space(r.get("제품명")) if "제품명" in arrange.columns else ""
            if v and k not in arrange_map:
                arrange_map[k] = v
            if nm and k not in arrange_name_map:
                arrange_name_map[k] = nm

    # Equipment line type: prefer '구분' column (line label). Fallback to blank.
    if "구분" in inj_equip.columns:
        inj_equip["라인구분"] = inj_equip["구분"].map(_norm_space)
    else:
        inj_equip["라인구분"] = ""

    name_to_base: dict[str, str] = {}
    name_key_to_base: dict[str, str] = {}
    if (not arrange.empty) and ("제품명" in arrange.columns) and ("제품명코드" in arrange.columns):
        tmp = arrange.loc[
            arrange["제품명"].astype("string").fillna("").astype(str).str.strip().ne(""),
            ["제품명", "제품명코드"],
        ].copy()

        def _name_key(v: object) -> str:
            s = _norm_space(v)
            if not s:
                return ""
            # Robust key: ignore spaces/underscores/dashes and most punctuation.
            s = s.lower()
            s = re.sub(r"[\s_\-]+", "", s)
            s = re.sub(r"[^0-9a-z가-힣]+", "", s)
            return s

        for _, r in tmp.iterrows():
            nm = _norm_space(r.get("제품명"))
            br = str(r.get("제품명코드") or "").strip().upper()
            if nm and br and nm not in name_to_base:
                name_to_base[nm] = br
            kk = _name_key(nm)
            if kk and br and kk not in name_key_to_base:
                name_key_to_base[kk] = br

    def _parse_running_base(v: object) -> str:
        s = str(v or "").strip()
        if not s:
            return ""
        # Prefer explicit R-code input in E column.
        if re.match(r"^R\d{3,}", s, flags=re.IGNORECASE):
            return _extract_base_r(s).upper()
        # Allowed fallback: map injection-sheet product name(E) -> arrange product name(J) -> base R(I).
        # (Do NOT use shortage-tab sales names for matching.)
        ns = _norm_space(s)
        out = str(name_to_base.get(ns, "")).strip().upper()
        if out:
            return out
        # More robust fuzzy match (handles minor punctuation/underscore differences).
        try:
            kk = re.sub(r"[\s_\-]+", "", ns.lower())
            kk = re.sub(r"[^0-9a-z가-힣]+", "", kk)
            return str(name_key_to_base.get(kk, "")).strip().upper()
        except Exception:
            return ""

    inj_equip["현재제품"] = inj_equip.get("생산 제품", "").map(_parse_running_base) if "생산 제품" in inj_equip.columns else ""
    inj_equip["설비명"] = inj_equip["설비코드"].astype("string").fillna("").astype(str).str.strip().str.upper()

    if "배정가능" in inj_equip.columns:
        usable = inj_equip.loc[inj_equip["배정가능"].fillna(False)].copy()
    else:
        usable = inj_equip.copy()
    usable = usable.loc[usable["설비명"].ne("")].copy()

    if usable.empty:
        warnings.append("사출 시트에서 배정 가능한 설비가 없습니다(E 공란 + F 기입 설비는 배정 불가).")
        return (pd.DataFrame(), pd.DataFrame(), warnings)

    if "라인구분" in usable.columns:
        blank_line = usable["라인구분"].map(_norm_space).eq("")
        if bool(blank_line.any()):
            examples = usable.loc[blank_line, "설비명"].astype("string").fillna("").astype(str).head(6).tolist()
            warnings.append(f"설비 라인구분(구분 컬럼) 인식 실패: {', '.join([e for e in examples if e])}")

    med_info = _load_injection_machine_medians_cached(excel_path, float(excel_mtime))
    med_by_machine: dict[str, float] = med_info.get("med_by_machine", {}) if isinstance(med_info, dict) else {}
    global_med = float(med_info.get("global_median", 0.0)) if isinstance(med_info, dict) else 0.0
    if global_med <= 0:
        global_med = 24000.0
        warnings.append("생산실적 기반 설비 생산량 추정에 실패하여 기본 CAPA(일 24000ea/설비)를 사용합니다.")

    def _equip_day_capa(equip_name: str) -> int:
        mc = _infer_machine_code_from_equip(equip_name)
        if mc and mc in med_by_machine:
            v = float(med_by_machine.get(mc) or 0.0)
            return int(max(0, round(v)))
        return int(max(0, round(global_med)))

    horizon_days = int(horizon_days)
    if horizon_days <= 0:
        horizon_days = 5
    days = [start_date + timedelta(days=i) for i in range(horizon_days)]

    work = demand.copy()
    for c in ["제품코드", "품명", "POWER", "납기일", "이니셜", "수주번호"]:
        if c not in work.columns:
            work[c] = ""
    if "사출" not in work.columns:
        work["사출"] = 0

    work["제품코드"] = work["제품코드"].astype("string").fillna("").astype(str).str.strip()
    work = work.loc[work["제품코드"].ne("")].copy()
    work = work.loc[work["제품코드"].str.upper().str.startswith("R")].copy()
    work["제품명코드"] = work["제품코드"].map(_extract_base_r).astype("string").fillna("").astype(str).str.strip().str.upper()
    work["사출"] = pd.to_numeric(work["사출"], errors="coerce").fillna(0).astype(int)
    work = work.loc[work["사출"].gt(0) & work["제품명코드"].ne("")].copy()
    if work.empty:
        return (pd.DataFrame(), pd.DataFrame(), ["R코드(사출) 부족수량(수요)이 없습니다."])

    work["POWER_num"] = pd.to_numeric(work["POWER"], errors="coerce")
    work["_due"] = work["납기일"].map(_coerce_date_value)

    product_info: dict[str, dict[str, object]] = {}

    def _order_ref(row: pd.Series) -> str:
        return _order_ref_string(row.get("이니셜"), row.get("수주번호"))

    for _, r in work.iterrows():
        base_r = str(r["제품명코드"] or "").strip().upper()
        if not base_r:
            continue
        due = r.get("_due", None)
        # IMPORTANT: shortage-tab '품명' is sales name; injection planning must use injection-sheet name(J) via mapping(I).
        name = str(arrange_name_map.get(base_r, "") or "").strip()
        p = r.get("POWER_num", None)
        if p is None or (isinstance(p, float) and math.isnan(p)):
            continue
        need = int(r.get("사출") or 0)
        if need <= 0:
            continue

        info = product_info.get(base_r)
        if info is None:
            info = {
                "제품명코드": base_r,
                "제품명": name,
                "납기일": due,
                "powers": {},
                "라인구분": arrange_map.get(base_r, ""),
                "order_refs": {},
            }
            product_info[base_r] = info
        if name and (not str(info.get("제품명") or "").strip()):
            info["제품명"] = name
        if due is not None:
            cur_due = info.get("납기일")
            if (cur_due is None) or (isinstance(cur_due, date) and due < cur_due):
                info["납기일"] = due

        # Track impacted orders (by initials/order no) for remaining view.
        ref = _order_ref(r)
        if ref:
            refs: dict[str, object] = info.get("order_refs") or {}
            prev = refs.get(ref)
            if prev is None:
                refs[ref] = due
            else:
                try:
                    pdue = _power_due_or_far(prev)
                    ndue = _power_due_or_far(due)
                    refs[ref] = ndue if ndue < pdue else pdue
                except Exception:
                    refs[ref] = prev or due
            info["order_refs"] = refs
        powers: dict[float, dict[str, object]] = info["powers"]  # type: ignore[assignment]
        pf = float(p)
        pinfo = powers.get(pf)
        if pinfo is None:
            pinfo = {"qty": 0, "due": due}
        # qty
        try:
            pinfo["qty"] = int(pinfo.get("qty") or 0) + int(need)
        except Exception:
            pinfo["qty"] = int(need)
        # due (min)
        if due is not None:
            cur_due = pinfo.get("due")
            if (cur_due is None) or (isinstance(cur_due, date) and isinstance(due, date) and due < cur_due):
                pinfo["due"] = due
            elif not isinstance(cur_due, date) and isinstance(due, date):
                pinfo["due"] = due
        powers[pf] = pinfo

    if not product_info:
        return (pd.DataFrame(), pd.DataFrame(), warnings or ["사출 수요가 없습니다."])

    def _product_remaining(base_r: str) -> int:
        info = product_info.get(base_r) or {}
        powers = info.get("powers") or {}
        out = 0
        for v in powers.values():
            try:
                out += int((v or {}).get("qty") or 0)
            except Exception:
                out += 0
        return int(out)

    def _product_due(base_r: str) -> date:
        d = product_info.get(base_r, {}).get("납기일")
        return d if isinstance(d, date) else date(2099, 12, 31)

    def _eligible_products_for_equipment(equip_row: pd.Series) -> list[str]:
        line = _norm_space(equip_row.get("라인구분"))
        if not line:
            return []
        out = [
            k
            for k, v in product_info.items()
            if _norm_space(v.get("라인구분")) == line and _product_remaining(k) > 0
        ]
        out.sort(key=lambda k: (_product_due(k), -_product_remaining(k), k))
        return out

    equip_last: dict[str, str] = {
        str(r["설비명"]): str(r.get("현재제품") or "").strip().upper() for _, r in usable.iterrows()
    }
    equip_affinity: dict[str, str] = {}
    equip_last_power: dict[str, float | None] = {str(r["설비명"]): None for _, r in usable.iterrows()}
    equip_last_power_set: dict[str, list[float]] = {str(r["설비명"]): [] for _, r in usable.iterrows()}

    rows: list[dict[str, object]] = []

    for day in days:
        # For "urgent" (due <= today/overdue) products, avoid pulling in additional machines unnecessarily.
        # Heuristic: if the remaining qty of an urgent product can be covered by machines that are *already*
        # running that product (yesterday/current), keep other machines on their current product to reduce churn.
        urgent_done_by_running: set[str] = set()
        try:
            # Remaining by product at the beginning of the day.
            urgent_remaining = {
                k: _product_remaining(k)
                for k in product_info.keys()
                if _product_remaining(k) > 0 and _product_due(k) <= day
            }
            if urgent_remaining:
                # Capacity by product from machines whose previous product is that product.
                cap_by_prod: dict[str, int] = {k: 0 for k in urgent_remaining.keys()}
                for _, er0 in usable.iterrows():
                    equip0 = str(er0.get("설비명") or "").strip().upper()
                    if not equip0:
                        continue
                    prev0 = str(equip_last.get(equip0, "") or "").strip().upper()
                    if prev0 and prev0 in cap_by_prod:
                        cap_by_prod[prev0] = int(cap_by_prod.get(prev0, 0)) + int(_equip_day_capa(equip0))
                for k, rem in urgent_remaining.items():
                    if int(cap_by_prod.get(k, 0)) >= int(rem):
                        urgent_done_by_running.add(k)
        except Exception:
            urgent_done_by_running = set()

        for _, er in usable.iterrows():
            equip_name = str(er.get("설비명") or "").strip().upper()
            if not equip_name:
                continue
            day_capa = _equip_day_capa(equip_name)
            block_capa = max(0, int(round(day_capa / 2.0)))
            # Operator rule: per shift, per cavity slot has a fixed max output.
            # (Overrides 생산실적 기반 CAPA-driven slot sizing.)
            slot_qty = 2000
            # Color matching / setup policy:
            # Avoid switching to a new product for tiny remainder in a block (causes matching churn).
            # If a new product cannot produce at least this qty in the block, leave the block idle unless urgent.
            min_new_product_qty = int(slot_qty)

            prev_prod = str(equip_last.get(equip_name, "") or "").strip().upper()
            affinity = str(equip_affinity.get(equip_name, "") or "").strip().upper()
            candidates = _eligible_products_for_equipment(er)
            prev_day_power = equip_last_power.get(equip_name, None)

            def _pick_product(prefer: str | None) -> str:
                if not candidates:
                    return ""
                best = candidates[0]
                best_due = _product_due(best)

                # Operational reality: for the *first planning day* (today),
                # keep the currently running product as-is when there is no "urgent" (due <= today) demand.
                # This avoids unnecessary back-and-forth assignments like "today A -> tomorrow B"
                # when the shop floor is already running B.
                prefer_u = str(prefer or "").strip().upper()
                if (
                    prefer_u
                    and (day == start_date)
                    and (best_due > day)
                    and (prefer_u in candidates)
                    and (_product_remaining(prefer_u) > 0)
                ):
                    return prefer_u

                # If the best (urgent) product can be fully handled by machines already running it,
                # don't switch other machines away from their current product just because it's overdue.
                if (
                    prefer_u
                    and (best_due <= day)
                    and (best in urgent_done_by_running)
                    and (prefer_u in candidates)
                    and (_product_remaining(prefer_u) > 0)
                ):
                    return prefer_u
                # Due date is the top priority. Keep previous/affinity only when they are within the same earliest due group.
                if prefer and prefer in candidates and _product_remaining(prefer) > 0 and _product_due(prefer) == best_due:
                    return prefer
                if affinity and affinity in candidates and _product_remaining(affinity) > 0 and _product_due(affinity) == best_due:
                    return affinity

                # Within earliest due group, *prefer* avoiding product switching for tiny remainder (unless urgent),
                # but do not block scheduling entirely: if any demand exists and the block would otherwise be idle,
                # we still need to assign the remaining qty somewhere.
                earliest = [k for k in candidates if _product_due(k) == best_due]
                if not earliest:
                    return best

                def _switch_penalty(k: str) -> int:
                    rem = int(_product_remaining(k) or 0)
                    due = _product_due(k)
                    urgent = bool(isinstance(due, date) and due <= day)
                    is_switch = bool(prefer_u and str(k).strip().upper() != prefer_u)
                    if is_switch and (not urgent) and rem < int(min_new_product_qty):
                        return 1
                    return 0

                # Sort by: (avoid tiny switch if possible) -> (larger remaining first)
                earliest.sort(key=lambda k: (_switch_penalty(k), -_product_remaining(k), k))
                return earliest[0]

            prod1 = _pick_product(prev_prod)
            if prod1:
                equip_affinity.setdefault(equip_name, prod1)

            def _setting_label(*, block: int, cur: str, prev_day: str, prev_block: str) -> str:
                cur = str(cur or "").strip().upper()
                prev_day = str(prev_day or "").strip().upper()
                prev_block = str(prev_block or "").strip().upper()
                if not cur:
                    return "유휴"
                if block == 1:
                    if not prev_day:
                        return "신규세팅"
                    return "세팅유지" if cur == prev_day else "잡체인지"
                if prev_block and (cur == prev_block):
                    return "세팅유지"
                if not prev_block and prev_day and (cur == prev_day):
                    return "세팅유지"
                if not prev_block and (not prev_day):
                    return "신규세팅"
                return "잡체인지"

            prev_block_prod = ""
            prev_block_power: float | None = None
            for block in (1, 2):
                if block == 1:
                    cur_prod = prod1
                else:
                    if prev_block_prod and _product_remaining(prev_block_prod) > 0:
                        cur_prod = prev_block_prod
                    else:
                        candidates = _eligible_products_for_equipment(er)
                        # Even when block1 product is exhausted, keep previous-day product as the "prefer" reference
                        # so we can avoid switching for tiny remainder (color matching churn).
                        cur_prod = _pick_product(prev_prod)
                        if cur_prod:
                            equip_affinity.setdefault(equip_name, cur_prod)

                powers_list: list[str] = []
                assign_qty = 0
                rem_after = 0
                prod_name = ""
                cur_primary_power: float | None = None
                prev_power_for_change = prev_day_power if block == 1 else prev_block_power
                chosen_types: list[float] = []
                if cur_prod:
                    info = product_info.get(cur_prod) or {}
                    prod_name = str(info.get("제품명") or "").strip()
                    powers: dict[float, dict[str, object]] = info.get("powers") or {}
                    prev_set = equip_last_power_set.get(equip_name, [])
                    slot_p, slot_q, chosen_types = _choose_power_slots_for_8cav(
                        powers,
                        prev_powers=prev_set,
                        block_day=day,
                        slots=8,
                        slot_qty=slot_qty,
                        max_types=8,
                    )
                    assign_qty = int(sum(slot_q))
                    rem_after = _product_remaining(cur_prod)
                    powers_list = [_fmt_power(p) for p in slot_p]
                    qty_list = [int(q) for q in slot_q]
                    cur_primary_power = slot_p[0] if slot_p else None
                    if len(powers_list) < 8:
                        powers_list += [""] * (8 - len(powers_list))
                    if len(qty_list) < 8:
                        qty_list += [0] * (8 - len(qty_list))
                else:
                    powers_list = [""] * 8
                    qty_list = [0] * 8

                # POWER variety / core change stats
                pw_seq = [p for p in powers_list if str(p).strip()]
                pw_unique = list(dict.fromkeys(pw_seq))
                pw_kind_count = int(len(set(pw_seq)))
                pw_change_in_block = 0
                for i in range(1, len(pw_seq)):
                    if pw_seq[i] != pw_seq[i - 1]:
                        pw_change_in_block += 1
                prev_pw_s = _fmt_power(prev_power_for_change) if isinstance(prev_power_for_change, float) else ""
                cur_pw_s = _fmt_power(cur_primary_power) if isinstance(cur_primary_power, float) else ""
                core_change = int(bool(prev_pw_s and cur_pw_s and (prev_pw_s != cur_pw_s)))

                setting = _setting_label(block=block, cur=cur_prod, prev_day=prev_prod, prev_block=prev_block_prod)
                row = {
                    "날짜": day,
                    "설비명": equip_name,
                    "제품명코드": cur_prod,
                    "제품명": prod_name,
                    "납기일": (product_info.get(cur_prod, {}).get("납기일") if cur_prod else None),
                    "Block": block,
                    "POWER 리스트": ", ".join([p for p in powers_list if str(p).strip()]),
                    "POWER 개수": int(sum(1 for p in powers_list if str(p).strip())),
                    "POWER(대표)": (pw_unique[0] if pw_unique else ""),
                    "POWER 종류수": pw_kind_count,
                    "POWER 변경횟수": int(pw_change_in_block),
                    "이전 POWER": prev_pw_s,
                    "코아교체": int(core_change),
                    "배정수량": int(assign_qty),
                    "잔여수량": int(rem_after) if cur_prod else 0,
                    "세팅구분": setting,
                }
                for i in range(8):
                    row[f"PW{i + 1}"] = str(powers_list[i] if i < len(powers_list) else "").strip()
                    row[f"Q{i + 1}"] = int(qty_list[i] if i < len(qty_list) else 0)
                rows.append(row)
                if cur_prod:
                    prev_block_prod = cur_prod
                    prev_block_power = cur_primary_power if isinstance(cur_primary_power, float) else prev_block_power
                    equip_last_power_set[equip_name] = [float(p) for p in slot_p if isinstance(p, (int, float))]

            equip_last[equip_name] = prev_block_prod or prev_prod
            equip_last_power[equip_name] = prev_block_power if isinstance(prev_block_power, float) else prev_day_power

    sched = pd.DataFrame(rows)

    all_lines = set(inj_equip["라인구분"].map(_norm_space).tolist()) if "라인구분" in inj_equip.columns else set()
    all_lines = {x for x in all_lines if x}
    usable_lines = set(usable["라인구분"].map(_norm_space).tolist()) if "라인구분" in usable.columns else set()
    usable_lines = {x for x in usable_lines if x}

    missing_arrange: list[str] = []
    rem_rows: list[dict[str, object]] = []

    def _format_refs(refs: dict[str, object], *, max_show: int = 10) -> str:
        items: list[tuple[date, str]] = []
        for k, v in (refs or {}).items():
            kk = str(k or "").strip()
            if not kk:
                continue
            items.append((_power_due_or_far(v), kk))
        items.sort(key=lambda t: (t[0], t[1]))
        keys = [k for _, k in items]
        if not keys:
            return ""
        if len(keys) <= int(max_show):
            return ", ".join(keys)
        return ", ".join(keys[: int(max_show)]) + f" …(+{len(keys) - int(max_show)})"

    for base_r, info in sorted(product_info.items(), key=lambda kv: (_product_due(kv[0]), kv[0])):
        rem = _product_remaining(base_r)
        if rem <= 0:
            continue
        line = _norm_space(info.get("라인구분"))
        if not line:
            reason = "어레인지 누락(I~K)"
            missing_arrange.append(base_r)
        elif line not in all_lines:
            reason = f"설비 라인 없음({line})"
        elif line not in usable_lines:
            reason = f"라인 설비 배정불가({line})"
        else:
            reason = "기간 CAPA 부족"
        rem_rows.append(
            {
                "제품명코드": base_r,
                "제품명": str(info.get("제품명") or "").strip(),
                "납기일": info.get("납기일"),
                "영향수주": _format_refs(info.get("order_refs") or {}),
                "잔여수량": rem,
                "미배정사유": reason,
            }
        )
    remaining = pd.DataFrame(rem_rows)

    if missing_arrange:
        missing_arrange = sorted(set(missing_arrange))
        previews = ", ".join(missing_arrange[:6])
        more = f" 외 {len(missing_arrange) - 6}개" if len(missing_arrange) > 6 else ""
        warnings.append(f"어레인지 누락({len(missing_arrange)}): {previews}{more} (사출 시트 I~K에 base R코드 매핑 필요)")
    return (sched, remaining, warnings)


@st.cache_data(show_spinner=False)
def _build_injection_schedule_cached(
    *,
    demand: pd.DataFrame,
    inj_equip: pd.DataFrame,
    arrange: pd.DataFrame,
    excel_path: str,
    excel_mtime: float,
    start_date: date,
    horizon_days: int,
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    return _build_injection_schedule(
        demand=demand,
        inj_equip=inj_equip,
        arrange=arrange,
        excel_path=excel_path,
        excel_mtime=excel_mtime,
        start_date=start_date,
        horizon_days=horizon_days,
    )


@st.cache_data(show_spinner=False)
def _build_injection_gantt_chart_df_cached(
    *,
    sched: pd.DataFrame,
    inj_equip: pd.DataFrame,
    start_date: date,
    horizon_days: int,
    now_block: int,
) -> tuple[pd.DataFrame, list[str]]:
    """
    Build gantt chart dataframe for injection schedule.
    Cached because it is UI-heavy and recomputed on every Streamlit rerun.
    """
    if sched is None or sched.empty:
        return (pd.DataFrame(), [])

    # Equipment list (include idle/disabled equipments for consistent rows)
    equip_all = inj_equip.copy()
    equip_all["설비명"] = (
        equip_all.get("설비코드", "")
        .astype("string")
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )
    equip_all["비고"] = equip_all.get("비고", "").astype("string").fillna("").astype(str).str.strip()
    equip_all["생산 제품"] = equip_all.get("생산 제품", "").astype("string").fillna("").astype(str).str.strip()
    equip_all["현재제품코드"] = equip_all.get("현재제품코드", "").astype("string").fillna("").astype(str).str.strip().str.upper()
    equip_all["배정가능"] = equip_all.get("배정가능", True).fillna(True)

    def _equip_sort_key(s: str) -> tuple[int, int, str]:
        s2 = str(s or "").strip().upper()
        m = re.match(r"^([A-Z])(\d+)$", s2)
        if not m:
            return (999, 999, s2)
        return (ord(m.group(1)) - 65, int(m.group(2)), s2)

    equip_list = sorted([e for e in equip_all["설비명"].tolist() if str(e).strip()], key=_equip_sort_key)
    equip_info = {
        str(r["설비명"]): {
            "배정가능": bool(r.get("배정가능", True)),
            "비고": str(r.get("비고") or "").strip(),
            "현재제품": str(r.get("생산 제품") or "").strip(),
            "현재제품코드": str(r.get("현재제품코드") or "").strip().upper(),
        }
        for _, r in equip_all.iterrows()
        if str(r.get("설비명") or "").strip()
    }

    s2 = sched.copy()
    s2["날짜"] = pd.to_datetime(s2["날짜"], errors="coerce").dt.date
    s2["Block"] = pd.to_numeric(s2["Block"], errors="coerce").fillna(0).astype(int)
    s2["설비명"] = s2["설비명"].astype("string").fillna("").astype(str).str.strip().str.upper()
    s2["제품명코드"] = s2["제품명코드"].astype("string").fillna("").astype(str).str.strip().str.upper()
    s2["제품명"] = s2["제품명"].astype("string").fillna("").astype(str).str.strip()
    s2["POWER 리스트"] = s2.get("POWER 리스트", "").astype("string").fillna("").astype(str).str.strip()
    s2["세팅구분"] = s2.get("세팅구분", "").astype("string").fillna("").astype(str).str.strip()
    s2["배정수량"] = pd.to_numeric(s2.get("배정수량", 0), errors="coerce").fillna(0).astype(int)
    s2["납기일"] = pd.to_datetime(s2.get("납기일", pd.NaT), errors="coerce").dt.strftime("%Y-%m-%d")

    slots = []
    horizon_days = int(horizon_days)
    for i in range(horizon_days):
        d = start_date + timedelta(days=i)
        for b, sh in [(1, "주간"), (2, "야간")]:
            slots.append({"날짜": d, "Block": b, "shift": sh, "slot_key": i * 2 + (0 if b == 1 else 1)})

    def _slot_label(d: date, sh: str) -> str:
        return f"{d.month}/{d.day} {sh}"

    s_map = {
        (r["설비명"], r["날짜"], int(r["Block"])): r.to_dict()
        for _, r in s2.iterrows()
        if isinstance(r.get("날짜"), date)
    }

    def _norm_r_code(v: object) -> str:
        s = str(v or "").strip().upper()
        if not s:
            return ""
        m = re.match(r"^(R\d+)", s)
        return m.group(1) if m else ""

    now_block = 2 if int(now_block) == 2 else 1

    # Precompute whether each equipment has any assigned product within the horizon.
    # Used to decide whether to "fill" the whole horizon with the currently running product
    # (only when schedule has no assignments at all for that equipment).
    assigned_any: dict[str, bool] = {str(e): False for e in equip_list}
    try:
        for (eq, _, _), rec in s_map.items():
            eqs = str(eq)
            if eqs not in assigned_any:
                continue
            pr = _norm_r_code((rec or {}).get("제품명코드"))
            if pr:
                assigned_any[eqs] = True
    except Exception:
        pass

    records: list[dict[str, object]] = []
    for e in equip_list:
        info = equip_info.get(str(e), {"배정가능": True, "비고": "", "현재제품": ""})
        e_key = str(e)
        has_any = bool(assigned_any.get(e_key, False))
        for sl in slots:
            d = sl["날짜"]
            b = int(sl["Block"])
            sh = str(sl["shift"])
            key = (str(e), d, b)
            src = s_map.get(key, None)
            prod_raw = str(src.get("제품명코드") or "").strip() if isinstance(src, dict) else ""
            prod = _norm_r_code(prod_raw)
            prod_name = str(src.get("제품명") or "").strip() if (isinstance(src, dict) and prod) else ""
            qty = int(src.get("배정수량") or 0) if isinstance(src, dict) else 0
            setting = str(src.get("세팅구분") or "").strip() if isinstance(src, dict) else ""
            pw = str(src.get("POWER 리스트") or "").strip() if isinstance(src, dict) else ""
            due = str(src.get("납기일") or "").strip() if isinstance(src, dict) else ""

            assignable = bool(info.get("배정가능", True))
            note = str(info.get("비고") or "").strip()
            cur_run = str(info.get("현재제품") or "").strip()
            cur_run_code = str(info.get("현재제품코드") or "").strip().upper()
            is_now_slot = isinstance(d, date) and (d == start_date) and (int(b) == int(now_block))
            fill_running_all = bool(cur_run_code) and (not has_any)

            if prod:
                state = "배정"
                idle_reason = ""
            else:
                qty = 0
                setting = ""
                pw = ""
                due = ""
                if not assignable:
                    state = "배정불가"
                    idle_reason = f"비고: {note}" if note else "비고: -"
                else:
                    state = "유휴"
                    idle_reason = ""

                # If schedule has no assignment for this equipment at all, keep the whole horizon occupied
                # by the currently running product (D2 같은 케이스). Still keep qty/power empty.
                if (state == "유휴") and fill_running_all:
                    state = "배정"
                    prod = cur_run_code
                    prod_name = cur_run
                    idle_reason = ""
                # Otherwise, show current running info only for the current slot (do not propagate to future slots).
                elif (state == "유휴") and is_now_slot and cur_run_code:
                    state = "배정"
                    prod = cur_run_code
                    prod_name = cur_run
                    idle_reason = ""
                elif (state == "유휴") and is_now_slot and cur_run:
                    idle_reason = f"현재 생산중: {cur_run}"

            records.append(
                {
                    "설비명": e,
                    "slot_label": _slot_label(d, sh),
                    "slot_key": int(sl["slot_key"]),
                    "주야": sh,
                    "상태": state,
                    "제품명코드": prod,
                    "제품명": prod_name,
                    "운영중제품": (cur_run if (is_now_slot and (not fill_running_all)) else ""),
                    "납기일": due,
                    "배정수량": qty,
                    "세팅구분": setting,
                    "POWER": pw,
                    "유휴사유": idle_reason,
                }
            )

    return (pd.DataFrame(records), equip_list)


def _injection_schedule_to_blocks(sched: pd.DataFrame) -> pd.DataFrame:
    """
    Convert schedule table into time blocks for gantt/grid rendering.
    Block 1 = 주간(08:00~20:00), Block 2 = 야간(20:00~익일 08:00).
    """
    if sched is None or sched.empty:
        return pd.DataFrame()
    df = sched.copy()
    if "날짜" not in df.columns or "Block" not in df.columns or "설비명" not in df.columns:
        return pd.DataFrame()

    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce").dt.date
    df["Block"] = pd.to_numeric(df["Block"], errors="coerce").fillna(0).astype(int)
    df["shift"] = df["Block"].map(lambda b: "주간" if int(b) == 1 else ("야간" if int(b) == 2 else ""))

    def _start_end(row: pd.Series) -> tuple[datetime | None, datetime | None]:
        d = row.get("날짜")
        if not isinstance(d, date):
            return (None, None)
        b = int(row.get("Block") or 0)
        if b == 1:
            s = datetime(d.year, d.month, d.day, 8, 0, 0)
            e = datetime(d.year, d.month, d.day, 20, 0, 0)
            return (s, e)
        if b == 2:
            s = datetime(d.year, d.month, d.day, 20, 0, 0)
            e = s + timedelta(hours=12)
            return (s, e)
        return (None, None)

    se = df.apply(_start_end, axis=1, result_type="expand")
    df["start"] = se[0]
    df["end"] = se[1]
    df = df.loc[df["start"].notna() & df["end"].notna()].copy()
    return df


@st.cache_data(show_spinner=False)
def _load_due_prepared(path: str, mtime: float) -> pd.DataFrame:
    # One-shot cache: load + prepare (avoids recomputing _prepare_lens_df on every rerun).
    base = _load_due_csv(path, mtime)
    return _prepare_lens_df(base)


@st.cache_data(show_spinner=False)
def _load_order_detail_prepared(path: str, mtime: float) -> pd.DataFrame:
    base = _load_order_detail_csv(path, mtime)
    out = _prepare_lens_df(base)
    # In order-detail, prefer the full product name if present.
    if "수요 제품 이름" in out.columns:
        out["품명"] = out["수요 제품 이름"].astype("string").fillna("")
    return out


@st.cache_data(show_spinner=False)
def _load_order_detail_grouped(path: str, mtime: float) -> pd.DataFrame:
    """
    Order view heavy step: group product-level rows once per source file.
    Filters (due-date/search/code) can be applied afterwards without regrouping.
    """
    df = _load_order_detail_prepared(path, mtime)
    numeric_cols = [c for c in [*DEFAULT_STAGE_COLS, "필요수량"] if c in df.columns]
    if numeric_cols:
        work = df.copy()
        for c in numeric_cols:
            work[c] = pd.to_numeric(work[c], errors="coerce").fillna(0)
    else:
        work = df

    group_cols = [c for c in ["이니셜", "수주번호", "신규분류 요약코드", "품명", "납기일"] if c in work.columns]
    if group_cols and numeric_cols:
        work = work.groupby(group_cols, dropna=False, as_index=False)[numeric_cols].sum(numeric_only=True)
    return work


def _prepare_lens_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "제품군" in out.columns:
        out["품명"], out["POWER"] = zip(*out["제품군"].map(_split_family))
    else:
        out["품명"] = out.get("수요 제품 이름", out.get("품명", "")).astype("string").fillna("")
        out["POWER"] = ""

    out["POWER"] = out["POWER"].astype("string")
    out["POWER_num"] = pd.to_numeric(out["POWER"], errors="coerce")
    power_num = out["POWER_num"]

    # POWER sort:
    # - Positive first: +10.00 -> +0.00
    # - Then zero: -0.00 (business display)
    # - Then negative: -0.00 -> -12.00 (magnitude increasing)
    out["POWER_group"] = 3
    out.loc[power_num.notna() & (power_num > 0), "POWER_group"] = 0
    out.loc[power_num.notna() & (power_num == 0), "POWER_group"] = 1
    out.loc[power_num.notna() & (power_num < 0), "POWER_group"] = 2
    out["POWER_sort"] = float("inf")
    out.loc[power_num.notna(), "POWER_sort"] = -power_num

    for c in ["ADD", "CP", "AXIS"]:
        if c in out.columns:
            out[c] = out[c].astype("string")

    # Lens type rules for display:
    # - Spherical: POWER only
    # - Toric: POWER + CP + AXIS
    # - Multifocal: POWER + ADD
    cp_raw = out["CP"] if "CP" in out.columns else pd.Series([""] * len(out), dtype="string")
    add_raw = out["ADD"] if "ADD" in out.columns else pd.Series([""] * len(out), dtype="string")
    axis_raw = out["AXIS"] if "AXIS" in out.columns else pd.Series([""] * len(out), dtype="string")

    cp_has = cp_raw.fillna("").astype(str).str.strip() != ""
    add_has = add_raw.fillna("").astype(str).str.strip() != ""
    axis_has = axis_raw.fillna("").astype(str).str.strip() != ""

    is_toric = cp_has | axis_has
    is_multi = (~is_toric) & add_has

    def _norm_axis(x: str) -> str:
        s = str(x).strip()
        if not s or s.lower() == "nan" or s == "<NA>":
            return "000"
        try:
            n = int(float(s))
            n = max(0, min(999, n))
            return f"{n:03d}"
        except Exception:
            return str(s).zfill(3)[:3]

    if "CP" in out.columns:
        out["CP"] = ""
        out.loc[is_toric, "CP"] = cp_raw.loc[is_toric].map(lambda x: _normalize_signed_2dp(x, zero_sign="+"))
        out.loc[is_toric & (out["CP"].astype(str).str.strip() == ""), "CP"] = "+0.00"
        out["CP"] = out["CP"].astype("string")
        out["CP_num"] = pd.to_numeric(out["CP"], errors="coerce").fillna(0)
    else:
        out["CP_num"] = 0

    if "AXIS" in out.columns:
        out["AXIS"] = ""
        out.loc[is_toric, "AXIS"] = axis_raw.loc[is_toric].map(_norm_axis)
        out.loc[is_toric & (out["AXIS"].astype(str).str.strip() == ""), "AXIS"] = "000"
        out["AXIS"] = out["AXIS"].astype("string")
        out["AXIS_num"] = pd.to_numeric(out["AXIS"], errors="coerce").fillna(0)
    else:
        out["AXIS_num"] = 0

    if "ADD" in out.columns:
        out["ADD"] = ""
        out.loc[is_multi, "ADD"] = add_raw.loc[is_multi].map(lambda x: _normalize_signed_2dp(x, zero_sign="+"))
        out.loc[is_multi & (out["ADD"].astype(str).str.strip() == ""), "ADD"] = "+0.00"
        out["ADD"] = out["ADD"].astype("string")
        out["ADD_num"] = pd.to_numeric(out["ADD"], errors="coerce").fillna(0)
    else:
        out["ADD_num"] = 0

    return out


def _apply_due_date_end_filter(df: pd.DataFrame, end: date) -> pd.DataFrame:
    if "납기일" not in df.columns:
        return df
    col = df["납기일"]
    due = col if pd.api.types.is_datetime64_any_dtype(col) else pd.to_datetime(col, errors="coerce")
    mask = due.dt.date.le(end)
    return df.loc[mask].copy()


def _compute_capa_table_from_prod_daily(
    prod_daily: pd.DataFrame,
    *,
    n_run_days: int,
    as_of: date,
) -> pd.DataFrame:
    """
    CAPA 산정: 최근 N개 '가동일'(실적>0) 기준 일평균 양품수량.
    - prod_daily는 (공정, 생산일자, 양품) 일별 집계 CSV 기준
    - as_of는 기준일(전일까지)로 자른다
    """
    if prod_daily is None or prod_daily.empty:
        return pd.DataFrame(columns=["공정", "CAPA", "capa_days", "last_run_date"])

    df = prod_daily
    if "생산일자" not in df.columns or "공정" not in df.columns or "양품" not in df.columns:
        return pd.DataFrame(columns=["공정", "CAPA", "capa_days", "last_run_date"])

    df = df.copy()
    df["생산일자"] = pd.to_datetime(df["생산일자"], errors="coerce")
    df["양품"] = pd.to_numeric(df["양품"], errors="coerce").fillna(0)
    df["공정"] = df["공정"].astype("string").fillna("").str.strip()

    # Filter once (avoid repeated .copy()).
    mask = df["공정"].ne("") & df["생산일자"].notna() & df["양품"].gt(0)
    if mask.any():
        df = df.loc[mask]
    else:
        return pd.DataFrame(columns=["공정", "CAPA", "capa_days", "last_run_date"])
    df = df.loc[df["생산일자"].dt.date.le(as_of)]
    if df.empty:
        return pd.DataFrame(columns=["공정", "CAPA", "capa_days", "last_run_date"])

    df = df.sort_values(["공정", "생산일자"], ascending=[True, True])

    rows: list[dict] = []
    for proc, g in df.groupby("공정", dropna=False):
        # 일별 집계 형태를 다시 보장(동일일자 다건 대비)
        by_day = (
            g.groupby(g["생산일자"].dt.date, dropna=False)["양품"]
            .sum()
            .sort_index()
        )
        if by_day.empty:
            continue
        last_days = by_day.tail(max(1, int(n_run_days)))
        rows.append(
            {
                "공정": str(proc),
                "CAPA": float(last_days.mean()),
                "capa_days": int(last_days.shape[0]),
                "last_run_date": pd.to_datetime(last_days.index[-1]),
            }
        )

    out = pd.DataFrame(rows)
    if not out.empty:
        out["CAPA"] = pd.to_numeric(out["CAPA"], errors="coerce").fillna(0)
    return out


def _grade_from_days(*, required_days: float, remaining_days: float, buffer_days: float) -> str:
    if required_days is None or remaining_days is None:
        return "NO_DUE"
    try:
        r = float(required_days)
        rem = float(remaining_days)
        buf = float(buffer_days)
    except Exception:
        return "NO_DUE"
    if rem < 0:
        rem = 0.0
    if r == 0.0:
        return "GREEN"
    if not pd.notna(r):
        return "NO_CAPA"
    if r == float("inf"):
        return "NO_CAPA"
    if r > rem:
        return "RED"
    if (rem - r) <= max(0.0, buf):
        return "YELLOW"
    return "GREEN"


def _build_order_risk_table(
    order_df: pd.DataFrame,
    capa_table: pd.DataFrame,
    *,
    today: date,
    buffer_days: float,
    start_offset_days: int = 1,
    injection_segs: list[dict[str, object]] | None = None,
    injection_start_date: date | None = None,
    injection_daily_fallback: float | None = None,
) -> pd.DataFrame:
    """
    수주별 리스크 산출.
    - 부족수량(order_df의 공정 컬럼) vs CAPA(공정별)
    - 수주 등급은 최악 공정 기준
    - 완료예상/지연은 누수규격 기준(연속 24/7 가정)
    """
    if order_df is None or order_df.empty:
        return pd.DataFrame()

    base = order_df.copy()
    if "납기일" in base.columns:
        due_col = base["납기일"]
        base["납기일"] = (
            due_col if pd.api.types.is_datetime64_any_dtype(due_col) else pd.to_datetime(due_col, errors="coerce")
        )

    stage_cols = [c for c in DEFAULT_STAGE_COLS if c in base.columns]
    if not stage_cols:
        return pd.DataFrame()

    for c in stage_cols:
        base[c] = pd.to_numeric(base[c], errors="coerce").fillna(0)

    # Remaining calendar days (24/7 운영 전제)
    due_dt = base["납기일"] if "납기일" in base.columns else pd.Series([pd.NaT] * len(base))
    base["_due_date"] = due_dt.dt.date
    base["남은일수_raw"] = base["_due_date"].map(lambda d: (d - today).days if isinstance(d, date) else None)
    base["남은일수"] = base["남은일수_raw"].map(lambda x: max(0, int(x)) if isinstance(x, int) else (max(0, int(x)) if isinstance(x, float) else None))

    capa_map = {}
    capa_days_map: dict[str, int] = {}
    if capa_table is not None and (not capa_table.empty) and "공정" in capa_table.columns:
        tmp = capa_table.copy()
        tmp["공정"] = tmp["공정"].astype("string").fillna("").str.strip()
        if "CAPA" in tmp.columns:
            tmp["CAPA"] = pd.to_numeric(tmp["CAPA"], errors="coerce").fillna(0)
        if "capa_days" in tmp.columns:
            tmp["capa_days"] = pd.to_numeric(tmp["capa_days"], errors="coerce").fillna(0).astype(int)
        capa_map = {str(r["공정"]): float(r.get("CAPA", 0) or 0) for r in tmp.to_dict("records")}
        capa_days_map = {str(r["공정"]): int(r.get("capa_days", 0) or 0) for r in tmp.to_dict("records")}

    meta_cols = [c for c in ["이니셜", "수주번호", "신규분류 요약코드", "품명", "납기일"] if c in base.columns]
    melt = base[meta_cols + ["남은일수_raw", "남은일수"] + stage_cols].melt(
        id_vars=meta_cols + ["남은일수_raw", "남은일수"],
        value_vars=stage_cols,
        var_name="공정",
        value_name="부족수량",
    )
    melt["부족수량"] = pd.to_numeric(melt["부족수량"], errors="coerce").fillna(0)
    melt = melt.loc[melt["부족수량"].gt(0)].copy()
    if melt.empty:
        return pd.DataFrame()

    melt["CAPA"] = melt["공정"].map(lambda x: capa_map.get(str(x), 0.0))
    melt["CAPA_days"] = melt["공정"].map(lambda x: capa_days_map.get(str(x), 0))
    melt["필요일수"] = melt.apply(
        lambda r: (float(r["부족수량"]) / float(r["CAPA"])) if float(r["CAPA"]) > 0 else float("inf"),
        axis=1,
    )
    melt["슬랙"] = melt.apply(
        lambda r: (float(r["남은일수"]) - float(r["필요일수"])) if r["남은일수"] is not None else float("nan"),
        axis=1,
    )
    melt["등급"] = melt.apply(
        lambda r: _grade_from_days(required_days=r["필요일수"], remaining_days=r["남은일수"], buffer_days=buffer_days),
        axis=1,
    )

    grade_rank = {"RED": 3, "YELLOW": 2, "GREEN": 1, "NO_CAPA": 4, "NO_DUE": 0}
    melt["_grade_rank"] = melt["등급"].map(lambda x: grade_rank.get(str(x), 0))

    group_key = [c for c in ["이니셜", "수주번호", "신규분류 요약코드"] if c in melt.columns]
    if not group_key:
        group_key = ["수주번호"] if "수주번호" in melt.columns else []
    if not group_key:
        return pd.DataFrame()

    # Worst grade per order: NO_CAPA > RED > YELLOW > GREEN
    worst = (
        melt.groupby(group_key, dropna=False)["_grade_rank"]
        .max()
        .reset_index()
        .rename(columns={"_grade_rank": "_worst_rank"})
    )

    proc_order = {"사출": 0, "분리": 1, "하이드레이션": 2, "접착": 3, "누수규격": 4}
    melt["_proc_order"] = melt["공정"].map(lambda x: proc_order.get(str(x), 999))

    # Start process: 가장 앞 공정(사출->...->누수규격) 중 부족수량>0인 공정
    idx_start = melt.groupby(group_key, dropna=False)["_proc_order"].idxmin()
    start_proc = melt.loc[idx_start, group_key + ["공정", "부족수량", "CAPA", "필요일수", "남은일수", "등급"]].copy()
    start_proc = start_proc.rename(
        columns={
            "공정": "시작공정",
            "부족수량": "시작_부족수량",
            "CAPA": "시작_CAPA",
            "필요일수": "시작_필요일수",
            "남은일수": "시작_남은일수",
            "등급": "시작_등급",
        }
    )

    # Bottleneck: (등급 심각도) -> (슬랙 최소) -> (공정순서) 우선
    def _pick_bn_index(g: pd.DataFrame) -> int:
        gg = g.sort_values(
            by=["_grade_rank", "슬랙", "_proc_order"],
            ascending=[False, True, True],
            na_position="last",
        )
        return int(gg.index[0])

    try:
        idx_bn = melt.groupby(group_key, dropna=False, group_keys=False).apply(_pick_bn_index, include_groups=False)
    except TypeError:
        # pandas<2.2 compatibility
        idx_bn = melt.groupby(group_key, dropna=False, group_keys=False).apply(_pick_bn_index)
    bottleneck = melt.loc[idx_bn, group_key + ["공정", "부족수량", "CAPA", "CAPA_days", "필요일수", "남은일수", "등급", "슬랙"]].copy()
    bottleneck = bottleneck.rename(
        columns={
            "공정": "병목공정",
            "부족수량": "병목_부족수량",
            "CAPA": "병목_CAPA",
            "CAPA_days": "병목_CAPA_days",
            "필요일수": "병목_필요일수",
            "남은일수": "병목_남은일수",
            "등급": "병목_등급",
            "슬랙": "병목_슬랙",
        }
    )

    # Join back representative meta columns
    rep = base.copy()
    rep_key = group_key.copy()
    rep_cols = [c for c in ["품명", "납기일"] if c in rep.columns]
    if rep_cols:
        rep = rep.groupby(rep_key, dropna=False, as_index=False).agg({c: "first" for c in rep_cols})
    else:
        rep = rep[group_key].drop_duplicates()

    out = (
        rep.merge(worst, on=group_key, how="left")
        .merge(start_proc, on=group_key, how="left")
        .merge(bottleneck, on=group_key, how="left")
    )

    def _rank_to_grade(n) -> str:
        try:
            n = int(n)
        except Exception:
            return "GREEN"
        inv = {v: k for k, v in grade_rank.items()}
        # Prefer business order: NO_CAPA as RED-equivalent alert
        g = inv.get(n, "GREEN")
        return "RED" if g == "NO_CAPA" else g

    out["리스크등급"] = out["_worst_rank"].map(_rank_to_grade)

    # Order-level shortage summary for risk display/forecast:
    # - 후공정 타관(사출/분리까지만 부족이고, 하이드/접착/누수는 0)
    # - 누수규격(표시용 필요수량): 기본은 누수규격 부족수량, 타관이면 분리(없으면 사출)
    # - 막힘공정(표시용): 공정간 부족 "증분"과 CAPA를 반영해 선택
    # - 완료예정일: 게이트공정 기준(타관이면 분리, 아니면 누수규격; 24/7 연속운영 가정)
    stage_agg = base[group_key + stage_cols].copy()
    stage_agg = stage_agg.groupby(group_key, dropna=False, as_index=False)[stage_cols].sum(numeric_only=True)
    for c in stage_cols:
        stage_agg[c] = pd.to_numeric(stage_agg[c], errors="coerce").fillna(0)
    s10 = stage_agg["사출"] if "사출" in stage_agg.columns else 0
    s20 = stage_agg["분리"] if "분리" in stage_agg.columns else 0
    s45 = stage_agg["하이드레이션"] if "하이드레이션" in stage_agg.columns else 0
    s55 = stage_agg["접착"] if "접착" in stage_agg.columns else 0
    s80 = stage_agg["누수규격"] if "누수규격" in stage_agg.columns else 0
    stage_agg["후공정_타관"] = ((s10 > 0) | (s20 > 0)) & (s45 <= 0) & (s55 <= 0) & (s80 <= 0)

    # 표시용 "누수규격"(필요수량)
    base_need = stage_agg["누수규격"] if "누수규격" in stage_agg.columns else 0
    alt_need = stage_agg["분리"] if "분리" in stage_agg.columns else (stage_agg["사출"] if "사출" in stage_agg.columns else 0)
    stage_agg["누수규격"] = base_need.where(~stage_agg["후공정_타관"], alt_need)

    # 부족 "증분" 계산(부족이 실제로 새로 생기는 구간)
    d10 = s10
    d20 = (s20 - s10).clip(lower=0)
    d45 = (s45 - s20).clip(lower=0)
    d55 = (s55 - s45).clip(lower=0)
    d80 = (s80 - s55).clip(lower=0)
    stage_agg["_d10"] = d10
    stage_agg["_d20"] = d20
    stage_agg["_d45"] = d45
    stage_agg["_d55"] = d55
    stage_agg["_d80"] = d80

    # 막힘공정 선택: (증분/CAPA) 최대 (CAPA=0이면 inf), 동률이면 증분 큰 공정, 그 다음 공정순서
    def _pick_block_stage(row) -> str:
        candidates = [
            ("사출", float(row.get("_d10") or 0)),
            ("분리", float(row.get("_d20") or 0)),
            ("하이드레이션", float(row.get("_d45") or 0)),
            ("접착", float(row.get("_d55") or 0)),
            ("누수규격", float(row.get("_d80") or 0)),
        ]
        best = ""
        best_score = -1.0
        best_delta = -1.0
        best_ord = 999
        for proc, delta in candidates:
            if delta <= 0:
                continue
            capa = float(capa_map.get(proc, 0.0) or 0.0)
            if capa <= 0:
                score = float("inf")
            else:
                score = delta / capa
            ordv = proc_order.get(proc, 999)
            if (
                (score > best_score)
                or (score == best_score and delta > best_delta)
                or (score == best_score and delta == best_delta and ordv < best_ord)
            ):
                best = proc
                best_score = score
                best_delta = delta
                best_ord = ordv
        # fallback: no incremental diff -> treat as first shortage stage
        if not best:
            for proc in DEFAULT_STAGE_COLS:
                try:
                    if float(row.get(proc, 0) or 0) > 0:
                        return proc
                except Exception:
                    continue
        return best

    stage_agg["막힘공정"] = stage_agg.apply(_pick_block_stage, axis=1)

    # 완료예정일(요청수량 기반 + 공정반영형):
    # - 공정별 필요수량(사출/분리/하이드/접착/누수)을 그대로 사용
    # - 공정 간 WIP/대기일은 공정당 1일로 고정
    # - 납기순(EDD)으로 수주를 정렬해 공정별 1라인(flow-shop) 스케줄링으로 완료예정일 산정
    due_agg = base[group_key + (["납기일"] if "납기일" in base.columns else [])].copy()
    if "납기일" in due_agg.columns:
        due_agg["납기일"] = pd.to_datetime(due_agg["납기일"], errors="coerce")
        due_agg = due_agg.groupby(group_key, dropna=False, as_index=False)["납기일"].min()
    else:
        due_agg = base[group_key].drop_duplicates()
        due_agg["납기일"] = pd.NaT

    stage_agg = stage_agg.merge(due_agg, on=group_key, how="left")

    # Gate stage for completion date: transfer -> 분리(없으면 사출), else 누수규격(없으면 마지막 부족 공정)
    def _gate_for_done(row) -> str:
        if bool(row.get("후공정_타관")):
            if "분리" in stage_cols and float(row.get("분리", 0) or 0) > 0:
                return "분리"
            if "사출" in stage_cols and float(row.get("사출", 0) or 0) > 0:
                return "사출"
        if "누수규격" in stage_cols and float(row.get("누수규격", 0) or 0) > 0:
            return "누수규격"
        last = ""
        last_ord = -1
        for p in DEFAULT_STAGE_COLS:
            if p in stage_cols and float(row.get(p, 0) or 0) > 0 and proc_order.get(p, -1) >= last_ord:
                last = p
                last_ord = proc_order.get(p, -1)
        return last

    stage_agg["게이트공정"] = stage_agg.apply(_gate_for_done, axis=1)

    if "납기일" in stage_agg.columns:
        due_col2 = stage_agg["납기일"]
        stage_agg["납기일"] = (
            due_col2 if pd.api.types.is_datetime64_any_dtype(due_col2) else pd.to_datetime(due_col2, errors="coerce")
        )

    sched = stage_agg.copy()
    sched = sched.sort_values(["납기일"], ascending=[True], na_position="last")

    proc_path_all = ["사출", "분리", "하이드레이션", "접착", "누수규격"]
    proc_path_all = [p for p in proc_path_all if p in stage_cols]

    completion: dict[str, list[float]] = {p: [] for p in proc_path_all}
    available: dict[str, float] = {p: 0.0 for p in proc_path_all}

    inj_allocator: _CapacityAllocator | None = None
    inj_start = injection_start_date if isinstance(injection_start_date, date) else today
    if injection_segs:
        segs = [dict(s) for s in injection_segs]
        # Fallback beyond horizon: extend segments with constant daily capacity until we cover all injection qty.
        try:
            total_need = float(sched.get("사출", 0).sum()) if "사출" in sched.columns else 0.0
        except Exception:
            total_need = 0.0
        planned_total = float(sum(float(s.get("cap") or 0) for s in segs))
        daily_fallback = float(injection_daily_fallback) if isinstance(injection_daily_fallback, (int, float)) else 0.0
        if daily_fallback <= 0:
            daily_fallback = float(capa_map.get("사출", 0.0) or 0.0)
        if daily_fallback > 0 and planned_total < total_need:
            rem = max(0.0, total_need - planned_total)
            per_block = max(0.0, daily_fallback / 2.0)
            # Extend at most 180 days to avoid runaway.
            ext_days = int(min(180, math.ceil(rem / max(1.0, daily_fallback))))
            # Continue from last segment end.
            last_t = float(segs[-1]["t0"]) + float(segs[-1]["dur"]) if segs else 0.0
            # Align to next day boundary.
            base_day = int(math.ceil(last_t))
            for di in range(ext_days):
                day0 = float(base_day + di)
                segs.append({"t0": day0 + 0.0, "dur": 0.5, "cap": int(round(per_block))})
                segs.append({"t0": day0 + 0.5, "dur": 0.5, "cap": int(round(per_block))})
        inj_allocator = _CapacityAllocator(segs)

    for _, row in sched.iterrows():
        is_transfer = bool(row.get("후공정_타관"))
        path = ["사출", "분리"] if is_transfer else proc_path_all
        prev_done = 0.0
        for p in proc_path_all:
            qty = float(row.get(p, 0) or 0)
            capa = float(capa_map.get(p, 0.0) or 0.0)
            # If process not in this order's path or has no qty, treat as no-op.
            if p not in path or qty <= 0:
                done = max(available[p], prev_done)
                completion[p].append(done)
                prev_done = done
                continue
            if p == "사출" and inj_allocator is not None:
                # Use injection plan capacity timeline (more realistic) instead of average CAPA.
                earliest = max(available[p], prev_done + float(RISK_WIP_DAYS_PER_PROCESS))
                done = inj_allocator.allocate(qty, earliest_start=float(earliest))
                completion[p].append(done)
                available[p] = max(available[p], done)
                prev_done = done
                continue
            if capa <= 0:
                done = float("inf")
                completion[p].append(done)
                prev_done = done
                available[p] = max(available[p], done)
                continue
            start = max(available[p], prev_done + float(RISK_WIP_DAYS_PER_PROCESS))
            dur = qty / capa
            done = start + dur
            completion[p].append(done)
            available[p] = done
            prev_done = done

    for p in proc_path_all:
        sched[f"_done_{p}"] = completion[p]

    def _done_day(row) -> float:
        gate = str(row.get("게이트공정") or "").strip()
        if not gate:
            return float("nan")
        try:
            return float(row.get(f"_done_{gate}", float("nan")))
        except Exception:
            return float("nan")

    sched["_done_days"] = sched.apply(_done_day, axis=1)

    def _to_date(x) -> pd.Timestamp:
        try:
            v = float(x)
        except Exception:
            return pd.NaT
        if not pd.notna(v) or v == float("inf"):
            return pd.NaT
        d = int(math.ceil(max(0.0, v + float(start_offset_days))))
        return pd.Timestamp(today + timedelta(days=d))

    sched["완료예정일"] = sched["_done_days"].map(_to_date)
    # NOTE: do not pre-create stage_agg["완료예정일"] before merge, otherwise pandas will suffix
    # columns (완료예정일_x/완료예정일_y) and downstream column selection will break.
    stage_agg = stage_agg.merge(sched[group_key + ["완료예정일"]], on=group_key, how="left")

    out = out.merge(
        stage_agg[group_key + ["후공정_타관", "누수규격", "막힘공정", "게이트공정", "완료예정일"]],
        on=group_key,
        how="left",
    )
    out["후공정_타관"] = out["후공정_타관"].fillna(False).astype(bool)
    if "누수규격" in out.columns:
        out["누수규격"] = pd.to_numeric(out["누수규격"], errors="coerce").fillna(0).astype(int)

    # 리스크가 없는(GREEN) 경우 완료예정일은 납기일로 표시(혼란 방지).
    if "납기일" in out.columns:
        out["납기일"] = pd.to_datetime(out["납기일"], errors="coerce")
    if "완료예정일" in out.columns:
        out["완료예정일"] = pd.to_datetime(out["완료예정일"], errors="coerce")
    if "리스크등급" in out.columns and "완료예정일" in out.columns and "납기일" in out.columns:
        mask_green = out["리스크등급"].astype(str).eq("GREEN") & out["납기일"].notna()
        out.loc[mask_green, "완료예정일"] = out.loc[mask_green, "납기일"]

    def _reason_row(r) -> str:
        if pd.isna(r.get("병목공정")) and pd.isna(r.get("시작공정")):
            return ""
        bn_proc_name = str(r.get("막힘공정") or r.get("병목공정") or "").strip()
        if not bn_proc_name:
            return ""

        bn_capa = float(capa_map.get(bn_proc_name, 0.0) or 0.0)
        grade = str(r.get("리스크등급") or "").strip()

        if grade == "GREEN":
            return ""

        if bn_capa <= 0:
            msg = f"{bn_proc_name} 병목(CAPA=0)"
        else:
            if grade == "RED":
                msg = f"{bn_proc_name} 병목(납기내 불가)"
            elif grade == "YELLOW":
                msg = f"{bn_proc_name} 주의(여유부족)"
            else:
                msg = f"{bn_proc_name} 정상"

        if bool(r.get("후공정_타관")):
            return f"{msg} / 후공정 타관"
        return msg

    out["리스크사유"] = out.apply(_reason_row, axis=1)

    # Priority sort: grade, due date, delay
    sort_grade = {"RED": 0, "YELLOW": 1, "GREEN": 2}
    out["_g"] = out["리스크등급"].map(lambda x: sort_grade.get(str(x), 9))
    if "납기일" in out.columns:
        out = out.sort_values(["_g", "납기일"], ascending=[True, True], na_position="last")
    else:
        out = out.sort_values(["_g"], ascending=[True], na_position="last")
    out.insert(0, "우선순위", range(1, len(out) + 1))
    out = out.drop(columns=[c for c in ["_g", "_worst_rank"] if c in out.columns])
    return out


@st.cache_data(show_spinner=False)
def _build_order_risk_table_cached(
    order_df: pd.DataFrame,
    capa_table: pd.DataFrame,
    *,
    today: date,
    buffer_days: float,
    start_offset_days: int = 1,
    injection_segs: list[dict[str, object]] | None = None,
    injection_start_date: date | None = None,
    injection_daily_fallback: float | None = None,
) -> pd.DataFrame:
    return _build_order_risk_table(
        order_df,
        capa_table,
        today=today,
        buffer_days=buffer_days,
        start_offset_days=start_offset_days,
        injection_segs=injection_segs,
        injection_start_date=injection_start_date,
        injection_daily_fallback=injection_daily_fallback,
    )


def main() -> None:
    st.title("S관 생산 필요수량 대시보드")
    _apply_local_theme_css()

    dashboard_links = _load_dashboard_links()
    with st.sidebar:
        if dashboard_links:
            st.markdown("<div class='sb-title'>대시보드 링크</div>", unsafe_allow_html=True)
            for item in dashboard_links:
                st.link_button(item["label"], item["url"], use_container_width=True)
            st.markdown("<div class='sb-hr'></div>", unsafe_allow_html=True)
        else:
            st.markdown("<div class='sb-title'>대시보드 링크</div>", unsafe_allow_html=True)
            st.caption(f"`{DASHBOARD_LINKS_PATH}`에 링크를 추가하면 여기서 새 탭으로 열 수 있어요.")
            st.markdown("<div class='sb-hr'></div>", unsafe_allow_html=True)

    excel_path = _find_repo_excel()
    if excel_path:
        st.caption(f"업데이트(APS raw data): `{_file_mtime_label(excel_path)}`")
    with st.sidebar:
        st.markdown("<div class='sb-title'>자료 다운로드</div>", unsafe_allow_html=True)
        b = _read_bytes(TEMPLATE_XLSX_PATH)
        if b is not None:
            st.download_button(
                "업로드 양식(.xlsx) 다운로드",
                data=b,
                file_name=os.path.basename(TEMPLATE_XLSX_PATH),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_tpl_xlsx",
            )
        else:
            st.caption(f"양식 파일이 없습니다: `{TEMPLATE_XLSX_PATH}`")

        if excel_path:
            st.caption(f"읽는 파일: `{os.path.basename(excel_path)}`")
        else:
            st.caption("읽는 파일: -")

    if not excel_path:
        st.error("`s관 부족수량.xlsx` 파일을 찾지 못했습니다. 저장소에 커밋해 두고 다시 실행하세요.")
        st.caption("기대 위치: `s관 부족수량.xlsx` 또는 `data/s관 부족수량.xlsx`")
        st.stop()

    out_dir = OUT_DIR
    status = _outputs_status(excel_path=excel_path, out_dir=out_dir)
    if not status.get("ok"):
        st.error(f"데이터 로딩 실패: {status.get('reason') or '-'}")
        st.stop()

    if status.get("needs_regen"):
        with st.spinner("엑셀 변경 감지: 데이터 생성 중..."):
            ensure = _ensure_latest_outputs(excel_path=excel_path, out_dir=out_dir)
    else:
        ensure = status

    if not ensure.get("ok"):
        st.error(f"데이터 생성 실패: {ensure.get('reason')}")
        st.stop()
    if ensure.get("regenerated"):
        st.cache_data.clear()

    due_csv = str(ensure["due_csv"])
    detail_csv = str(ensure["detail_csv"])
    equip_code_target_csv = ensure.get("equip_code_target_csv")
    prod_daily_csv = ensure.get("prod_daily_csv")

    detail_for_map: pd.DataFrame | None = None
    try:
        detail_for_map = _load_order_detail_prepared(detail_csv, os.path.getmtime(detail_csv))
    except Exception:
        detail_for_map = None

    equip_code_target_df: pd.DataFrame | None = None
    try:
        if equip_code_target_csv:
            equip_code_target_df = _load_equip_code_min_target_csv(
                str(equip_code_target_csv), os.path.getmtime(str(equip_code_target_csv))
            )
        else:
            equip_code_target_df = None
    except Exception:
        equip_code_target_df = None

    def _tab_sort_key(code: str) -> tuple[int, int, str]:
        s = str(code).strip()
        sl = s.lower()
        is_color = 1 if "color" in sl else 0  # non-color first, color later
        if "toric" in sl:
            lens_rank = 2
        elif "m/f" in sl or "multifocal" in sl or "multi" in sl:
            lens_rank = 1
        elif "sph" in sl or "spherical" in sl:
            lens_rank = 0
        else:
            lens_rank = 9
        return (is_color, lens_rank, sl)

    try:
        base = _load_due_prepared(due_csv, os.path.getmtime(due_csv))
    except Exception as e:
        st.error("데이터 로딩 실패")
        st.caption(str(e))
        st.stop()

    df = base
    new_code_col = "신규분류 요약코드" if "신규분류 요약코드" in df.columns else None

    def render(
        filtered: pd.DataFrame,
        *,
        ui_key_prefix: str,
        process_only: str | None = None,
        selected_code: str | None = None,
    ) -> None:
        stage_cols_raw = [process_only] if process_only else DEFAULT_STAGE_COLS
        numeric_cols = [c for c in stage_cols_raw if c in filtered.columns]
        total_label = f"{process_only} 필요수량" if process_only else "총 필요수량"
        header_ph = st.empty()
        metric_ph = st.empty()

        search_raw = st.text_input(
            "검색 (품명)",
            placeholder="예: O2O2, SEPIA, ASH",
            key=f"{ui_key_prefix}_name_search",
        )
        filtered = _filter_by_name_contains(filtered, "품명", search_raw)

        if process_only and process_only in filtered.columns:
            proc_v = pd.to_numeric(filtered[process_only], errors="coerce").fillna(0)
            filtered = filtered.loc[proc_v.ne(0)].copy()

        df_num = filtered.copy()
        for c in numeric_cols:
            df_num[c] = pd.to_numeric(df_num[c], errors="coerce").fillna(0)

        total_col = process_only if process_only in df_num.columns else ("누수규격" if "누수규격" in df_num.columns else None)
        header_ph.subheader(total_label)
        metric_ph.metric(label="", value=_format_int(df_num[total_col].sum()) if total_col else "0")

        st.subheader("납기별 상세")

        view = filtered.copy()
        sort_cols: list[str] = []
        for c in [
            "납기일",
            "최소목표일",
            "신규분류 요약코드",
            "품명",
            "POWER_group",
            "POWER_sort",
            "CP_num",
            "AXIS_num",
            "ADD_num",
        ]:
            if c in view.columns:
                sort_cols.append(c)
        if sort_cols:
            view = view.sort_values(sort_cols, ascending=[True] * len(sort_cols), na_position="last")

        allowed_prefixes: list[str] | None = None
        if process_only:
            prefix_map: dict[str, list[str]] = {
                "사출": ["R"],
                "분리": ["Q"],
                "하이드레이션": ["P"],
                "접착": ["P"],
                "누수규격": ["P"],
            }
            allowed_prefixes = prefix_map.get(process_only, None)
            view = _attach_item_codes(view, detail_for_map, allowed_prefixes=allowed_prefixes)
            if "제품코드" not in view.columns:
                view["제품코드"] = ""

            if (
                detail_for_map is not None
                and equip_code_target_df is not None
                and (not equip_code_target_df.empty)
                and "제품 코드" in detail_for_map.columns
                and "최소목표일" in equip_code_target_df.columns
                and "제품 코드" in equip_code_target_df.columns
                and "공정" in equip_code_target_df.columns
            ):
                d = detail_for_map.copy()
                if "납기일" in d.columns:
                    d["납기일"] = pd.to_datetime(d["납기일"], errors="coerce")
                d["제품 코드"] = d["제품 코드"].astype("string").fillna("").str.strip()
                if allowed_prefixes:
                    prefixes = [str(p).strip().upper() for p in allowed_prefixes if str(p).strip()]
                    if prefixes:
                        d = d.loc[d["제품 코드"].map(lambda x: any(str(x).upper().startswith(p) for p in prefixes))]
                t = equip_code_target_df.loc[equip_code_target_df["공정"].astype("string") == process_only, ["제품 코드", "최소목표일"]]
                d = d.merge(t, on="제품 코드", how="left")
                key_candidates = ["신규분류 요약코드", "제품군", "납기일"]
                key_cols = [c for c in key_candidates if c in view.columns and c in d.columns]
                if key_cols:
                    tgt = d.groupby(key_cols, dropna=False, as_index=False).agg(최소목표일=("최소목표일", "min"))
                    view = view.merge(tgt, on=key_cols, how="left")
            if "최소목표일" not in view.columns:
                view["최소목표일"] = pd.NaT

        # Export dataframe (keep numeric) BEFORE display formatting.
        export_df = view.copy()
        for s in stage_cols_raw:
            if s in export_df.columns:
                export_df[s] = pd.to_numeric(export_df[s], errors="coerce").fillna(0).astype(int)

        has_toric = False
        has_multi = False
        if "CP" in view.columns or "AXIS" in view.columns:
            cp_has = (
                view["CP"].astype("string").fillna("").astype(str).str.strip().ne("").any()
                if "CP" in view.columns
                else False
            )
            axis_has = (
                view["AXIS"].astype("string").fillna("").astype(str).str.strip().ne("").any()
                if "AXIS" in view.columns
                else False
            )
            has_toric = bool(cp_has or axis_has)
        if "ADD" in view.columns:
            has_multi = bool(view["ADD"].astype("string").fillna("").astype(str).str.strip().ne("").any())

        # 공정별 보기에서 분류가 명확한 경우 규격 컬럼을 강제 결정
        if selected_code and selected_code != "전체":
            sl = str(selected_code).lower()
            if "toric" in sl:
                has_toric, has_multi = True, False
            elif "m/f" in sl or "multifocal" in sl or "multi" in sl:
                has_toric, has_multi = False, True
            elif "sph" in sl or "spherical" in sl:
                has_toric, has_multi = False, False

        cols = ["신규분류 요약코드", "품명"]
        if process_only:
            cols.append("제품코드")
        cols += ["POWER"]
        cols += ["납기일"]
        if process_only:
            cols.append("최소목표일")
        cols += stage_cols_raw
        if has_toric:
            power_idx = cols.index("POWER")
            cols[power_idx + 1 : power_idx + 1] = ["CP", "AXIS"]
        if has_multi:
            power_idx = cols.index("POWER")
            insert_at = power_idx + 1 + (2 if has_toric else 0)
            cols[insert_at:insert_at] = ["ADD"]

        cols = [c for c in cols if c in view.columns]

        export_cols = cols.copy()
        export_df2 = export_df.copy()

        # Download rules:
        # - 납기별 상세: 화면 그대로(제품코드 없음)
        # - 공정별 보기: 제품코드 포함 + 공정별 prefix 필터
        if process_only:
            if "제품코드" in export_df2.columns and "제품코드" not in export_cols:
                # 품명과 POWER 사이에 제품코드
                if "품명" in export_cols and "POWER" in export_cols:
                    export_cols.insert(export_cols.index("POWER"), "제품코드")
                elif "품명" in export_cols:
                    export_cols.insert(export_cols.index("품명") + 1, "제품코드")
                else:
                    export_cols.insert(0, "제품코드")

        # Display formatting (comma) after export snapshot.
        for s in stage_cols_raw:
            if s in view.columns:
                view[s] = pd.to_numeric(view[s], errors="coerce").fillna(0).astype(int)

        column_config = {
            "신규분류 요약코드": st.column_config.TextColumn(width="medium"),
            "품명": st.column_config.TextColumn(width="large"),
            "제품코드": st.column_config.TextColumn(width="medium"),
            "POWER": st.column_config.TextColumn(width="small"),
            "CP": st.column_config.TextColumn(width="small"),
            "AXIS": st.column_config.TextColumn(width="small"),
            "ADD": st.column_config.TextColumn(width="small"),
            "최소목표일": st.column_config.DatetimeColumn(format="YYYY-MM-DD", width="small"),
            "납기일": st.column_config.DatetimeColumn(format="YYYY-MM-DD", width="small"),
        }
        for c in stage_cols_raw:
            if c in cols:
                column_config[c] = st.column_config.NumberColumn(format="localized", width="small")
        column_config = {k: v for k, v in column_config.items() if k in cols}

        # Download button should not push the totals row away from the table header,
        # so render it BEFORE totals grid.
        xlsx_bytes = _to_excel_bytes(export_df2[export_cols], sheet_name="다운로드")
        st.download_button(
            "엑셀 다운로드",
            data=xlsx_bytes,
            file_name=f"{'공정' if process_only else '납기'}_{selected_code or '전체'}_{process_only or '전체'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{ui_key_prefix}_download",
        )

        stage_totals = {c: _format_int(df_num[c].sum()) for c in stage_cols_raw if c in df_num.columns}
        view_show = view[cols].copy()
        for c in stage_cols_raw:
            if c in view_show.columns:
                view_show[c] = pd.to_numeric(view_show[c], errors="coerce").fillna(0).map(_format_int)
        # Display totals as a clean HTML row above the columns
        if stage_totals:
            totals_html = " ".join([
                f"<span style='margin-right: 20px; font-size: 15px;'>{c}: <strong style='color: #0066cc;'>{stage_totals[c]}</strong></span>"
                for c in stage_cols_raw if c in stage_totals
            ])
            st.markdown(f"<div style='margin-bottom: 8px; padding: 4px 8px;'>{totals_html}</div>", unsafe_allow_html=True)

        view_show.columns = cols

        table_h = _table_height_for_rows(len(view), min_height=280, max_height=720)
        st.dataframe(
            _style_dataframe_like_dashboard(view_show),
            use_container_width=True,
            height=table_h,
            hide_index=True,
            column_config=column_config,
        )

    if new_code_col is None:
        render(df, ui_key_prefix="all")
        return

    view_options = ["납기별 상세", "공정별 보기", "수주별 현황", "리스크", "사출 계획"]
    _pre_widget_single_select_fix(key="view_mode", default="납기별 상세", options=view_options)
    view_mode_raw = st.segmented_control(
        "보기",
        options=view_options,
        default="납기별 상세",
        key="view_mode",
        on_change=_on_change_single_select,
        args=("view_mode", "납기별 상세", view_options),
        label_visibility="collapsed",
    )
    view_mode = _coerce_single_value(view_mode_raw, default="납기별 상세", options=view_options)

    prev_mode = st.session_state.get("_prev_view_mode")
    if prev_mode != view_mode:
        st.session_state["code_pill"] = ["전체"]
        if view_mode == "납기별 상세":
            # Reset due-date filter when entering due view.
            st.session_state["due_due_quick"] = "해제"
            st.session_state["due_due_end"] = _today_kst()
            st.session_state["_prev_due_due_quick"] = "해제"
        if view_mode == "공정별 보기":
            st.session_state["process_pill"] = "사출"
            # Reset due-date filter when entering process view.
            st.session_state["proc_due_quick"] = "해제"
            st.session_state["proc_due_end"] = _today_kst()
            st.session_state["_prev_proc_due_quick"] = "해제"
        if view_mode == "수주별 현황":
            # Always reset due-date filter when entering order view.
            st.session_state["order_due_quick"] = "해제"
            st.session_state["order_due_end"] = _today_kst()
            st.session_state["_prev_order_due_quick"] = "해제"
        if view_mode == "리스크":
            st.session_state["risk_due_quick"] = "해제"
            st.session_state["risk_due_end"] = _today_kst()
            st.session_state["_prev_risk_due_quick"] = "해제"
            st.session_state["risk_grade_pill"] = ["RED", "YELLOW"]
        st.session_state["_prev_view_mode"] = view_mode

    process_only = None
    if view_mode == "납기별 상세":
        # Due date end quick-picks for due view.
        due_quick_options = ["해제", "직접", "당월", "+7일", "+14일"]
        _pre_widget_single_select_fix(key="due_due_quick", default="해제", options=due_quick_options)
        due_quick_raw = st.pills(
            "납기일 종료 (빠른 선택)",
            options=due_quick_options,
            default="해제",
            key="due_due_quick",
            selection_mode="single",
            on_change=_on_change_single_select,
            args=("due_due_quick", "해제", due_quick_options),
            label_visibility="collapsed",
        )
        due_quick = _coerce_single_value(due_quick_raw, default="해제", options=due_quick_options)
        if due_quick == "당월":
            due_default_end = _end_of_month(_today_kst())
        elif due_quick == "+7일":
            due_default_end = _today_kst() + timedelta(days=7)
        elif due_quick == "+14일":
            due_default_end = _today_kst() + timedelta(days=14)
        else:
            due_default_end = _today_kst()

        prev_due_quick = st.session_state.get("_prev_due_due_quick")
        if prev_due_quick != due_quick:
            st.session_state["due_due_end"] = due_default_end
            st.session_state["_prev_due_due_quick"] = due_quick

        due_end_date = st.date_input(
            "납기일 종료",
            value=st.session_state.get("due_due_end", due_default_end),
            key="due_due_end",
            disabled=(due_quick == "해제"),
        )

    if view_mode == "공정별 보기":
        _pre_widget_single_select_fix(key="process_pill", default="사출", options=DEFAULT_STAGE_COLS)
        process_only_raw = st.pills(
            "공정",
            options=DEFAULT_STAGE_COLS,
            default="사출",
            key="process_pill",
            on_change=_on_change_single_select,
            args=("process_pill", "사출", DEFAULT_STAGE_COLS),
            label_visibility="collapsed",
        )
        process_only = _coerce_single_value(process_only_raw, default="사출", options=DEFAULT_STAGE_COLS)

        # Due date end quick-picks (same idea as order view).
        proc_quick_options = ["해제", "직접", "당월", "+7일", "+14일"]
        _pre_widget_single_select_fix(key="proc_due_quick", default="해제", options=proc_quick_options)
        proc_quick_raw = st.pills(
            "납기일 종료 (빠른 선택)",
            options=proc_quick_options,
            default="해제",
            key="proc_due_quick",
            selection_mode="single",
            on_change=_on_change_single_select,
            args=("proc_due_quick", "해제", proc_quick_options),
            label_visibility="collapsed",
        )
        proc_quick = _coerce_single_value(proc_quick_raw, default="해제", options=proc_quick_options)
        if proc_quick == "당월":
            proc_default_end = _end_of_month(_today_kst())
        elif proc_quick == "+7일":
            proc_default_end = _today_kst() + timedelta(days=7)
        elif proc_quick == "+14일":
            proc_default_end = _today_kst() + timedelta(days=14)
        else:
            proc_default_end = _today_kst()

        prev_proc_quick = st.session_state.get("_prev_proc_due_quick")
        if prev_proc_quick != proc_quick:
            st.session_state["proc_due_end"] = proc_default_end
            st.session_state["_prev_proc_due_quick"] = proc_quick

        proc_end_date = st.date_input(
            "납기일 종료",
            value=st.session_state.get("proc_due_end", proc_default_end),
            key="proc_due_end",
            disabled=(proc_quick == "해제"),
        )

    # NOTE: 사출 계획은 항상 "현재 기준 5일 계획"을 보여주며,
    # 납기일 종료 필터에 의해 계획이 바뀌는 구조가 아니므로 별도 필터 UI를 노출하지 않는다.

    # 분류 pills (view-mode별로 totals 계산 데이터가 다름)
    order_df_all: pd.DataFrame | None = None
    if view_mode in ("수주별 현황", "리스크"):
        if detail_for_map is None:
            st.error("수주별/리스크 데이터가 없습니다. 엑셀 시트/컬럼을 확인하세요.")
            st.stop()
        order_df = _load_order_detail_grouped(detail_csv, os.path.getmtime(detail_csv))
        if view_mode == "리스크":
            # 리스크의 완료예정일(스케줄)은 전체 backlog 기준으로 계산해야 하므로,
            # 표시 필터(납기일 종료) 적용 전 원본을 별도로 보관한다.
            order_df_all = order_df.copy()

        # Due date end quick-picks
        quick_key = "order_due_quick" if view_mode == "수주별 현황" else "risk_due_quick"
        prev_quick_key = "_prev_order_due_quick" if view_mode == "수주별 현황" else "_prev_risk_due_quick"
        end_key = "order_due_end" if view_mode == "수주별 현황" else "risk_due_end"

        quick_options = ["해제", "직접", "당월", "+7일", "+14일"]
        default_quick = "해제"
        _pre_widget_single_select_fix(key=quick_key, default=default_quick, options=quick_options)
        quick_raw = st.pills(
            "납기일 종료 (빠른 선택)",
            options=quick_options,
            default=default_quick,
            key=quick_key,
            selection_mode="single",
            on_change=_on_change_single_select,
            args=(quick_key, default_quick, quick_options),
            label_visibility="collapsed",
        )
        quick = _coerce_single_value(quick_raw, default=default_quick, options=quick_options)
        if quick == "당월":
            default_end = _end_of_month(_today_kst())
        elif quick == "+7일":
            default_end = _today_kst() + timedelta(days=7)
        elif quick == "+14일":
            default_end = _today_kst() + timedelta(days=14)
        else:
            default_end = _today_kst()

        # Ensure quick pick actually updates the date_input (Streamlit keeps widget state by key).
        prev_quick = st.session_state.get(prev_quick_key)
        if prev_quick != quick:
            st.session_state[end_key] = default_end
            st.session_state[prev_quick_key] = quick

        end_date = st.date_input(
            "납기일 종료",
            value=st.session_state.get(end_key, default_end),
            key=end_key,
            disabled=(quick == "해제"),
        )
        if quick != "해제":
            order_df = _apply_due_date_end_filter(order_df, end_date)
        codes_src = order_df
        value_col = "누수규격"
    else:
        codes_src = df
        if view_mode == "납기별 상세":
            due_quick_state = st.session_state.get("due_due_quick", "해제")
            if due_quick_state != "해제":
                codes_src = _apply_due_date_end_filter(
                    codes_src,
                    st.session_state.get("due_due_end", _today_kst()),
                )
        if view_mode == "공정별 보기":
            proc_quick_state = st.session_state.get("proc_due_quick", "해제")
            if proc_quick_state != "해제":
                codes_src = _apply_due_date_end_filter(codes_src, st.session_state.get("proc_due_end", _today_kst()))
        if view_mode == "사출 계획":
            value_col = "사출" if "사출" in codes_src.columns else "누수규격"
        else:
            value_col = process_only if process_only else "누수규격"

    totals_base: dict[str, float] = {}
    total_all = 0.0
    if value_col in codes_src.columns:
        tmp = codes_src.copy()
        tmp[value_col] = pd.to_numeric(tmp[value_col], errors="coerce").fillna(0)
        totals_base = tmp.groupby(new_code_col, dropna=False)[value_col].sum(numeric_only=True).to_dict()
        total_all = float(tmp[value_col].sum())

    codes = (
        codes_src[new_code_col]
        .astype("string")
        .fillna("")
        .map(lambda x: x.strip() if isinstance(x, str) else "")
    )
    code_options = sorted([c for c in codes.unique().tolist() if c], key=_tab_sort_key)

    def _code_label(code: str) -> str:
        if code == "전체":
            return f"전체 ({_format_int(total_all)})"
        return f"{code} ({_format_int(totals_base.get(code, 0.0))})"

    code_all_options = ["전체"] + code_options

    # For injection plan, show these as *informational chips* (not clickable filters).
    # Many users naturally click pills because they look interactive.
    if view_mode == "사출 계획":
        chips = []
        for c in code_all_options:
            chips.append(
                f"""<span class="aps-chip"><span class="aps-chip-label">{c}</span><span class="aps-chip-val">{_format_int(total_all) if c=='전체' else _format_int(totals_base.get(c, 0.0))}</span></span>"""
            )
        st.markdown(
            """
<style>
.aps-chip-wrap { display: flex; flex-wrap: wrap; gap: 6px 8px; margin: 2px 0 10px 0; }
.aps-chip {
  display: inline-flex;
  align-items: center;
  gap: 8px;
  padding: 6px 10px;
  border: 1px solid rgba(0,0,0,0.14);
  border-radius: 999px;
  background: rgba(0,0,0,0.03);
  font-size: 13px;
  line-height: 1.1;
  cursor: default;
  user-select: none;
}
.aps-chip-label { font-weight: 600; }
.aps-chip-val { font-weight: 800; }
</style>
            """,
            unsafe_allow_html=True,
        )
        st.markdown(f"<div class='aps-chip-wrap'>{''.join(chips)}</div>", unsafe_allow_html=True)
        codes_selected = ["전체"]
    else:
        _pre_widget_multi_select_fix(key="code_pill", default=["전체"], options=code_all_options)
        code_raw = st.pills(
            "분류",
            options=code_all_options,
            default=["전체"],
            key="code_pill",
            format_func=_code_label,
            selection_mode="multi",
            on_change=_on_change_multi_select_all_exclusive,
            args=("code_pill", "전체", code_all_options),
            label_visibility="collapsed",
        )
        codes_selected = _coerce_multi_values(code_raw, default=["전체"], options=code_all_options)

    def _is_all_codes(v: list[str]) -> bool:
        return (not v) or ("전체" in v)

    def _codes_label(v: list[str]) -> str:
        if _is_all_codes(v):
            return "전체"
        if len(v) == 1:
            return v[0]
        vv = [str(x).strip() for x in v if str(x).strip()]
        return f"{vv[0]}+{len(vv) - 1}"

    def _codes_key(v: list[str]) -> str:
        if _is_all_codes(v):
            return "전체"
        try:
            import hashlib

            s = "|".join(sorted(set([str(x).strip() for x in v if str(x).strip()])))
            return "M_" + hashlib.md5(s.encode("utf-8")).hexdigest()[:8]
        except Exception:
            return "M_MULTI"

    code_label = _codes_label(codes_selected)
    code_key = _codes_key(codes_selected)

    if view_mode == "수주별 현황":
        subset = (
            order_df
            if _is_all_codes(codes_selected)
            else order_df[order_df[new_code_col].astype("string").isin(codes_selected)].copy()
        )
        stage_cols_raw = DEFAULT_STAGE_COLS
        numeric_cols = [c for c in stage_cols_raw if c in subset.columns]

        search_raw = st.text_input(
            "검색 (품명/이니셜/수주번호)",
            placeholder="예: 해외, 202601, SEPIA",
            key=f"order_{code_key}_search",
        )
        subset = _filter_by_any_contains(subset, ["품명", "이니셜", "수주번호"], search_raw)

        detail_num = subset.copy()
        for c in numeric_cols:
            if c in detail_num.columns:
                detail_num[c] = pd.to_numeric(detail_num[c], errors="coerce").fillna(0)

        stage_sum = 0
        for c in numeric_cols:
            stage_sum = stage_sum + detail_num[c].fillna(0)
        detail_num = detail_num.loc[stage_sum.fillna(0).gt(0)].copy()

        stage_totals = {
            c: _format_int(pd.to_numeric(detail_num[c], errors="coerce").fillna(0).sum()) for c in numeric_cols
        }

        # Summary rows: order-level (initial + order number), summed across products.
        summary_base = detail_num.copy()
        if "납기일" in summary_base.columns:
            due_dt = pd.to_datetime(summary_base["납기일"], errors="coerce")
            summary_base["_due_date"] = due_dt.dt.date
        else:
            summary_base["_due_date"] = pd.NaT

        group_key = [c for c in ["이니셜", "수주번호"] if c in summary_base.columns]
        if not group_key:
            group_key = ["수주번호"] if "수주번호" in summary_base.columns else []
        if (_is_all_codes(codes_selected) or len(codes_selected) > 1) and new_code_col in summary_base.columns:
            group_key = [c for c in [*group_key, new_code_col] if c in summary_base.columns]

        agg_spec: dict[str, str] = {c: "sum" for c in numeric_cols}
        if "품명" in summary_base.columns:
            agg_spec["품목수"] = "nunique"
        else:
            agg_spec["품목수"] = "size"
        if "_due_date" in summary_base.columns:
            agg_spec["납기일"] = "min"

        # Pandas needs existing columns for named aggs; create working cols.
        work = summary_base.copy()
        if "품명" in work.columns:
            work["품목수"] = work["품명"]
        else:
            work["품목수"] = 1
        if "_due_date" in work.columns:
            work["납기일"] = work["_due_date"]

        order_num = work.groupby(group_key, dropna=False, as_index=False).agg(agg_spec)
        sort_cols = [c for c in ["납기일", new_code_col, "이니셜", "수주번호"] if c in order_num.columns]
        if sort_cols:
            order_num = order_num.sort_values(sort_cols, ascending=[True] * len(sort_cols), na_position="last")
        order_num.insert(0, "우선순위", range(1, len(order_num) + 1))

        order_view = order_num.copy()
        for c in numeric_cols:
            if c in order_view.columns:
                order_view[c] = pd.to_numeric(order_view[c], errors="coerce").fillna(0).astype(int)
        if "품목수" in order_view.columns:
            order_view["품목수"] = pd.to_numeric(order_view["품목수"], errors="coerce").fillna(0).astype(int)

        base_cols = ["우선순위", "이니셜", "수주번호"]
        if _is_all_codes(codes_selected) or len(codes_selected) > 1:
            base_cols.append(new_code_col)
        base_cols += ["품목수", "납기일"]
        summary_cols = [c for c in base_cols if c in order_view.columns] + numeric_cols

        col_cfg_summary: dict[str, object] = {
            "우선순위": st.column_config.NumberColumn(format="%d", width="small"),
            "이니셜": st.column_config.TextColumn(width="small"),
            "수주번호": st.column_config.TextColumn(width="medium"),
            "신규분류 요약코드": st.column_config.TextColumn(width="medium"),
            "품목수": st.column_config.NumberColumn(format="%d", width="small"),
            "납기일": st.column_config.DatetimeColumn(format="YYYY-MM-DD", width="small"),
        }
        for c in numeric_cols:
            col_cfg_summary[c] = st.column_config.NumberColumn(format="localized", width="small")
        col_cfg_summary = {k: v for k, v in col_cfg_summary.items() if k in summary_cols}

        xlsx_bytes_sum = _to_excel_bytes(order_view[summary_cols], sheet_name="수주요약")
        st.download_button(
            "엑셀 다운로드 (요약)",
            data=xlsx_bytes_sum,
            file_name=f"수주요약_{code_label}_{_today_kst().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"order_{code_key}_download_sum",
        )

        order_show = order_view[summary_cols].copy()
        order_show.columns = summary_cols

        # Display totals as HTML above the table
        if stage_totals:
            totals_html = " ".join([
                f"<span style='margin-right: 20px; font-size: 15px;'>{c}: <strong style='color: #0066cc;'>{stage_totals[c]}</strong></span>"
                for c in numeric_cols if c in stage_totals
            ])
            st.markdown(f"<div style='margin-bottom: 8px; padding: 4px 8px;'>{totals_html}</div>", unsafe_allow_html=True)

        sum_h = _table_height_for_rows(len(order_view), min_height=260, max_height=520)
        st.dataframe(
            _style_dataframe_like_dashboard(order_show),
            use_container_width=True,
            height=sum_h,
            hide_index=True,
            column_config=col_cfg_summary,
        )

        st.divider()

        view = detail_num.copy()
        sort_cols = [c for c in ["납기일", "이니셜", "수주번호", "품명"] if c in view.columns]
        if sort_cols:
            view = view.sort_values(sort_cols, ascending=[True] * len(sort_cols), na_position="last")
        view.insert(0, "우선순위", range(1, len(view) + 1))

        for c in numeric_cols:
            if c in view.columns:
                view[c] = pd.to_numeric(view[c], errors="coerce").fillna(0).astype(int)

        cols = [c for c in ["우선순위", "이니셜", "수주번호", "신규분류 요약코드", "품명", "납기일"] if c in view.columns] + numeric_cols

        column_config = {
            "우선순위": st.column_config.NumberColumn(format="%d", width="small"),
            "이니셜": st.column_config.TextColumn(width="small"),
            "수주번호": st.column_config.TextColumn(width="medium"),
            "신규분류 요약코드": st.column_config.TextColumn(width="medium"),
            "품명": st.column_config.TextColumn(width="large"),
            "납기일": st.column_config.DatetimeColumn(format="YYYY-MM-DD", width="small"),
        }
        for c in numeric_cols:
            column_config[c] = st.column_config.NumberColumn(format="localized", width="small")
        column_config = {k: v for k, v in column_config.items() if k in cols}

        xlsx_bytes_det = _to_excel_bytes(view[cols], sheet_name="수주상세")
        st.download_button(
            "엑셀 다운로드 (상세)",
            data=xlsx_bytes_det,
            file_name=f"수주상세_{code_label}_{_today_kst().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"order_{code_key}_download_det",
        )

        detail_h = _table_height_for_rows(len(view), min_height=320, max_height=720)
        st.dataframe(
            _style_dataframe_like_dashboard(view[cols]),
            use_container_width=True,
            height=detail_h,
            hide_index=True,
            column_config=column_config,
        )
        return

    if view_mode == "리스크":
        st.subheader("리스크 (수주 기준)")

        if not prod_daily_csv:
            st.error("`생산실적` 시트 기반 CAPA 데이터가 없습니다. 엑셀에 `생산실적` 시트가 있는지 확인하세요.")
            st.stop()

        st.caption(
            f"기준: CAPA=최근 {RISK_CAPA_RUN_DAYS} 가동일(전일까지, 양품>0) 평균 · "
            f"YELLOW 버퍼={RISK_YELLOW_BUFFER_DAYS:.0f}일 · 24/7 연속운영 가정"
        )
        st.caption(
            "등급: RED=납기내 불가(필요일수>남은일수) · YELLOW=여유부족(버퍼 1일 이내) · GREEN=가능"
        )

        # 완료예정일(스케줄)은 전체 backlog 기준으로 계산하고, 화면 표시만 필터 적용한다.
        schedule_src = order_df_all if isinstance(order_df_all, pd.DataFrame) and not order_df_all.empty else order_df
        view_src = order_df

        def _to_order_level(df0: pd.DataFrame) -> pd.DataFrame:
            stage_cols = [c for c in [*DEFAULT_STAGE_COLS, "필요수량"] if c in df0.columns]
            if not stage_cols:
                return df0
            work = df0.copy()
            for c in stage_cols:
                work[c] = pd.to_numeric(work[c], errors="coerce").fillna(0)
            if "납기일" in work.columns:
                work["납기일"] = pd.to_datetime(work["납기일"], errors="coerce")
            group_key = [c for c in ["이니셜", "수주번호", "신규분류 요약코드"] if c in work.columns]
            if not group_key:
                group_key = ["수주번호"] if "수주번호" in work.columns else []
            if not group_key:
                return work
            agg: dict[str, str] = {c: "sum" for c in stage_cols}
            if "납기일" in work.columns:
                agg["납기일"] = "min"
            if "품명" in work.columns:
                agg["품명"] = "first"
            return work.groupby(group_key, dropna=False, as_index=False).agg(agg)

        schedule_orders = _to_order_level(schedule_src)
        view_orders = _to_order_level(view_src)

        prod_daily_df = _load_prod_daily_csv(str(prod_daily_csv), os.path.getmtime(str(prod_daily_csv)))
        as_of = _today_kst() - timedelta(days=1)
        capa_table = _compute_capa_table_from_prod_daily(prod_daily_df, n_run_days=int(RISK_CAPA_RUN_DAYS), as_of=as_of)

        # Optional: reflect injection short-term schedule into completion dates (more realistic injection start/throughput).
        inj_excel_mtime = float(os.path.getmtime(excel_path)) if excel_path and os.path.exists(excel_path) else 0.0
        inj_info = _load_injection_sheet_cached(excel_path, inj_excel_mtime) if excel_path else {"equip": pd.DataFrame(), "arrange": pd.DataFrame()}
        inj_equip = inj_info.get("equip", pd.DataFrame())
        inj_arrange = inj_info.get("arrange", pd.DataFrame())
        inj_segs: list[dict[str, object]] = []
        inj_daily_fallback = 0.0
        try:
            if capa_table is not None and (not capa_table.empty) and ("공정" in capa_table.columns) and ("CAPA" in capa_table.columns):
                mask_inj = capa_table["공정"].astype("string").fillna("").astype(str).str.strip().eq("사출")
                if bool(mask_inj.any()):
                    inj_daily_fallback = float(pd.to_numeric(capa_table.loc[mask_inj, "CAPA"].head(1), errors="coerce").fillna(0).iloc[0])
        except Exception:
            inj_daily_fallback = 0.0
        try:
            inj_demand = df.copy()
            # Reduce columns early (cache hashing + schedule work).
            keep_cols = [c for c in ["제품코드", "품명", "POWER", "납기일", "사출", "이니셜", "수주번호"] if c in inj_demand.columns]
            inj_demand = inj_demand[keep_cols].copy() if keep_cols else inj_demand
            inj_demand = _attach_item_codes(inj_demand, detail_for_map, allowed_prefixes=["R"])
            if "제품코드" not in inj_demand.columns:
                inj_demand["제품코드"] = ""
            if "사출" not in inj_demand.columns:
                inj_demand["사출"] = 0
            inj_demand["사출"] = pd.to_numeric(inj_demand["사출"], errors="coerce").fillna(0).astype(int)
            inj_demand = inj_demand.loc[inj_demand["사출"].gt(0)].copy()
            if (not inj_demand.empty) and (inj_equip is not None) and (not inj_equip.empty):
                inj_segs = _build_injection_plan_segments_cached(
                    demand=inj_demand,
                    inj_equip=inj_equip,
                    inj_arrange=inj_arrange,
                    excel_path=excel_path,
                    excel_mtime=inj_excel_mtime,
                    start_date=_today_kst(),
                    horizon_days=5,
                )
        except Exception:
            inj_segs = []

        # Compute risk/schedule on ALL orders (schedule_orders), then filter down for display (subset).
        risk_all = _build_order_risk_table_cached(
            schedule_orders,
            capa_table,
            today=_today_kst(),
            buffer_days=float(RISK_YELLOW_BUFFER_DAYS),
            start_offset_days=int(RISK_SCHED_START_OFFSET_DAYS),
            injection_segs=inj_segs if inj_segs else None,
            injection_start_date=_today_kst(),
            injection_daily_fallback=inj_daily_fallback,
        )
        if risk_all.empty:
            st.caption("표시할 리스크 대상이 없습니다.")
            st.stop()

        # Base filters (code + due) for counts/pills.
        risk_base = risk_all.copy()
        if (not _is_all_codes(codes_selected)) and new_code_col in risk_base.columns:
            risk_base = risk_base.loc[risk_base[new_code_col].astype("string").isin(codes_selected)].copy()
        if quick != "해제" and "납기일" in risk_base.columns:
            risk_base = _apply_due_date_end_filter(risk_base, st.session_state.get("risk_due_end", _today_kst()))

        grade_options = ["RED", "YELLOW", "GREEN"]
        counts = risk_base["리스크등급"].value_counts().to_dict() if "리스크등급" in risk_base.columns else {}
        filter_options = grade_options

        def _grade_label(opt: str) -> str:
            s = str(opt)
            if s in grade_options:
                return f"{s} ({int(counts.get(s, 0))})"
            return s

        grade_raw = st.pills(
            "등급 필터",
            options=filter_options,
            default=st.session_state.get("risk_grade_pill", ["RED", "YELLOW"]),
            key="risk_grade_pill",
            format_func=_grade_label,
            selection_mode="multi",
            on_change=_on_change_risk_grade_pills,
            kwargs={"key": "risk_grade_pill", "grade_options": grade_options},
            label_visibility="collapsed",
        )
        selected_grades = grade_raw if isinstance(grade_raw, list) else ([grade_raw] if grade_raw else [])
        selected_grades = [g for g in selected_grades if g in grade_options]
        if not selected_grades:
            selected_grades = ["RED", "YELLOW"]

        search_raw = st.text_input(
            "검색 (이니셜/수주번호/품명)",
            placeholder="예: 해외, 202601, O2O2",
            key=f"risk_{code_key}_search",
        )

        risk_df = _filter_by_any_contains(risk_base, ["품명", "이니셜", "수주번호"], search_raw)
        risk_df = risk_df.loc[risk_df["리스크등급"].isin(selected_grades)].copy()
        if risk_df.empty:
            st.caption("필터 조건에 해당하는 항목이 없습니다.")
            st.stop()
        st.caption(f"표시 건수: {len(risk_df):,}")

        show_cols = [
            "우선순위",
            "이니셜",
            "수주번호",
            "신규분류 요약코드",
            "품명",
            "납기일",
            "리스크등급",
            "리스크사유",
            "누수규격",
            "완료예정일",
        ]
        show_cols = [c for c in show_cols if c in risk_df.columns]

        xlsx_bytes_risk = _to_excel_bytes(risk_df[show_cols], sheet_name="리스크")
        st.download_button(
            "엑셀 다운로드",
            data=xlsx_bytes_risk,
            file_name=f"리스크_{code_label}_{_today_kst().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"risk_{code_key}_download",
        )

        column_config = {
            "우선순위": st.column_config.NumberColumn(format="%d", width="small"),
            "이니셜": st.column_config.TextColumn(width="small"),
            "수주번호": st.column_config.TextColumn(width="medium"),
            "신규분류 요약코드": st.column_config.TextColumn(width="medium"),
            "품명": st.column_config.TextColumn(width="large"),
            "납기일": st.column_config.DatetimeColumn(format="YYYY-MM-DD", width="small"),
            "리스크등급": st.column_config.TextColumn(width="small"),
            "리스크사유": st.column_config.TextColumn(width="large"),
            "누수규격": st.column_config.NumberColumn(format="localized", width="small"),
            "완료예정일": st.column_config.DatetimeColumn(format="YYYY-MM-DD", width="small"),
        }
        column_config = {k: v for k, v in column_config.items() if k in show_cols}

        table_h = _table_height_for_rows(len(risk_df), min_height=320, max_height=780)
        st.dataframe(
            _style_dataframe_like_dashboard(risk_df[show_cols]),
            use_container_width=True,
            height=table_h,
            hide_index=True,
            column_config=column_config,
        )
        return

    if view_mode == "사출 계획":
        st.subheader("사출 스케줄 (자동 생성)")
        # Fixed horizon + start date (planning is always shown from 'today').
        horizon_days = 5
        start_date = _today_kst()
        now_block = 2 if datetime.now(KST).hour >= 20 else 1

        # Keep "간트" as default; guard old persisted selection values (e.g. "그리드").
        inj_view_options = ["간트", "상세표"]
        if st.session_state.get("inj_view_kind") not in inj_view_options:
            st.session_state["inj_view_kind"] = "간트"
        view_kind = st.segmented_control(
            "표시",
            options=inj_view_options,
            default="간트",
            key="inj_view_kind",
            label_visibility="collapsed",
        )

        excel_mtime = float(os.path.getmtime(excel_path))
        base_df = df

        # NOTE: product-group pills are for *volume visibility* only.
        # Injection schedule must always be generated from the full dataset (전체),
        # otherwise users change the plan by just clicking a pill.
        demand_for_sched = base_df.copy()
        demand_for_sched = _attach_item_codes(demand_for_sched, detail_for_map, allowed_prefixes=["R"])
        if "제품코드" not in demand_for_sched.columns:
            demand_for_sched["제품코드"] = ""

        inj_info = _load_injection_sheet_cached(excel_path, excel_mtime)
        inj_equip = inj_info.get("equip", pd.DataFrame())
        inj_arrange = inj_info.get("arrange", pd.DataFrame())
        if inj_equip is None or inj_equip.empty:
            st.error("엑셀에 `사출` 시트(설비 현황)가 없습니다.")
            st.stop()

        sched, remaining, warns = _build_injection_schedule_cached(
            demand=demand_for_sched,
            inj_equip=inj_equip,
            arrange=inj_arrange,
            excel_path=excel_path,
            excel_mtime=excel_mtime,
            start_date=start_date,
            horizon_days=horizon_days,
        )

        if warns:
            for w in warns[:8]:
                st.warning(w)

        if sched is None or sched.empty:
            st.caption("생성된 스케줄이 없습니다.")
            return

        xlsx_ops = _to_injection_operation_xlsx_cached(
            sched,
            start_date=start_date,
            horizon_days=horizon_days,
            sheet_name="운영양식",
            equip_all=inj_equip,
            excel_mtime=excel_mtime,
            now_block=now_block,
        )
        st.download_button(
            "엑셀 다운로드 (상세표)",
            data=xlsx_ops,
            file_name=f"사출상세표_{code_label}_{_today_kst().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"inj_{code_key}_download_ops",
        )

        blocks = _injection_schedule_to_blocks(sched)
        has_assigned = bool(blocks.get("제품명코드", pd.Series(dtype=str)).astype("string").fillna("").str.strip().ne("").any())
        if not has_assigned:
            st.info("현재 조건에서 배정 가능한 제품 후보가 없어 모든 블록이 유휴로 표시됩니다. (어레인지/라인구분/배정불가 설비를 확인하세요)")

        if view_kind == "간트":
            # Render as a day/shift grid (주간/야간) instead of detailed time axis.
            chart_df, equip_list = _build_injection_gantt_chart_df_cached(
                sched=sched,
                inj_equip=inj_equip,
                start_date=start_date,
                horizon_days=horizon_days,
                now_block=now_block,
            )
            if chart_df is None or chart_df.empty or not equip_list:
                st.caption("간트 표시할 데이터가 없습니다.")
            else:
                # Legend label: show code + product name (many users don't memorize R-codes).
                def _norm_name(s: str) -> str:
                    return " ".join(str(s or "").split()).strip()

                chart_df["제품라벨"] = None
                if (not chart_df.empty) and ("제품명코드" in chart_df.columns) and ("제품명" in chart_df.columns):
                    c = chart_df["제품명코드"].astype("string").fillna("").astype(str).str.strip().str.upper()
                    n = chart_df["제품명"].astype("string").fillna("").astype(str).str.strip()
                    lab = c
                    has_name = n.str.strip().ne("")
                    # Use multi-line label: "R코드\n품명" (keep full name for readability).
                    lab = lab.where(~has_name, c + "\n" + n.map(_norm_name))
                    # Only assigned rows should appear in product legend (avoid "null").
                    if "상태" in chart_df.columns:
                        is_assigned = chart_df["상태"].astype("string").fillna("").astype(str).str.strip().eq("배정")
                        lab = lab.where(is_assigned, None)
                    chart_df["제품라벨"] = lab.where(c.str.strip().ne(""), None)

                legend_values: list[str] = []
                try:
                    tmp = chart_df.loc[
                        chart_df.get("상태", "").astype("string").fillna("").astype(str).str.strip().eq("배정")
                        & chart_df.get("제품라벨", "").astype("string").fillna("").astype(str).str.strip().ne(""),
                        ["제품명코드", "제품라벨"],
                    ].copy()
                    if not tmp.empty:
                        tmp["제품명코드"] = tmp["제품명코드"].astype("string").fillna("").astype(str).str.strip().str.upper()
                        tmp["제품라벨"] = tmp["제품라벨"].astype("string").fillna("").astype(str)
                        tmp = tmp.drop_duplicates(subset=["제품명코드"], keep="first").sort_values("제품명코드", ascending=True)
                        legend_values = tmp["제품라벨"].tolist()
                except Exception:
                    legend_values = []

                spec = {
                    "$schema": "https://vega.github.io/schema/vega-lite/v5.json",
                    "data": {"values": chart_df.to_dict(orient="records")},
                    "encoding": {
                        "y": {"field": "설비명", "type": "nominal", "sort": equip_list, "axis": {"title": ""}},
                        "x": {
                            "field": "slot_label",
                            "type": "ordinal",
                            "sort": {"field": "slot_key", "op": "min"},
                            "axis": {"title": "", "labelAngle": 0},
                        },
                        "tooltip": [
                            {"field": "설비명", "type": "nominal", "title": "호기"},
                            {"field": "slot_label", "type": "nominal", "title": "일자/주야"},
                            {"field": "상태", "type": "nominal", "title": "상태"},
                            {"field": "제품명코드", "type": "nominal", "title": "R코드"},
                            {"field": "제품명", "type": "nominal", "title": "품명(사출시트)"},
                            {"field": "운영중제품", "type": "nominal", "title": "현재제품(설비)"},
                            {"field": "배정수량", "type": "quantitative", "title": "배정수량"},
                            {"field": "세팅구분", "type": "nominal", "title": "세팅"},
                            {"field": "유휴사유", "type": "nominal", "title": "유휴/사유"},
                        ],
                    },
                    "layer": [
                        {
                            "mark": {"type": "rect", "stroke": "#dcdcdc", "strokeWidth": 1},
                            "encoding": {
                                "color": {
                                    "condition": [
                                        {"test": "datum.상태 === '배정불가'", "value": "#d0d0d0"},
                                        {"test": "datum.상태 === '유휴'", "value": "#f2f2f2"},
                                    ],
                                    "field": "제품라벨",
                                    "type": "nominal",
                                    "scale": {"scheme": "tableau20"},
                                    "legend": {
                                        "title": "제품",
                                        "values": legend_values if legend_values else [],
                                        "labelLimit": 520,
                                        "labelFontSize": 12,
                                        "titleFontSize": 13,
                                        "symbolSize": 90,
                                        "labelLineHeight": 16,
                                    },
                                }
                            },
                        },
                        {
                            "mark": {"type": "text", "baseline": "middle", "align": "center", "fontSize": 12},
                            "encoding": {
                                "text": {"condition": {"test": "datum.상태 === '배정'", "field": "제품명코드"}, "value": ""},
                                "color": {
                                    "condition": [{"test": "datum.상태 === '배정'", "value": "#111111"}],
                                    "value": "#666666",
                                },
                            },
                        },
                    ],
                    "config": {
                        "axis": {
                            "grid": True,
                            "gridColor": "#e5e5e5",
                            "gridOpacity": 1,
                            "domain": False,
                            "labelFontSize": 12,
                            "titleFontSize": 12,
                        },
                        "view": {"stroke": "transparent"},
                    },
                    "height": max(380, min(980, 26 * int(len(equip_list) + 1))),
                }

                st.vega_lite_chart(spec, use_container_width=True)

                # Optional: daily summary (useful for quick demand/assignment check), keep collapsed by default.
                with st.expander("일자별 세부 타겟(요약)", expanded=False):
                    tgt = chart_df.loc[chart_df["상태"].eq("배정")].copy()
                    if tgt.empty:
                        st.caption("집계할 배정 데이터가 없습니다.")
                    else:
                        daily = tgt.groupby(["slot_label", "제품명코드", "제품명", "납기일"], dropna=False, as_index=False).agg(
                            블록수=("배정수량", "size"),
                            배정수량=("배정수량", "sum"),
                        )
                        daily["배정수량"] = pd.to_numeric(daily["배정수량"], errors="coerce").fillna(0).astype(int)
                        daily = daily.sort_values(["slot_label", "납기일", "배정수량"], ascending=[True, True, False], na_position="last")
                        st.dataframe(
                            _style_dataframe_like_dashboard(daily),
                            use_container_width=True,
                            height=_table_height_for_rows(len(daily), min_height=220, max_height=520),
                            hide_index=True,
                        )
        else:
            # 상세표(운영양식 형태): 모든 설비를 표시 (유휴/배정불가 설비 포함)
            view_df = sched.copy() if isinstance(sched, pd.DataFrame) else pd.DataFrame()
            if "설비명" in view_df.columns:
                view_df["설비명"] = view_df["설비명"].astype("string").fillna("").astype(str).str.strip().str.upper()
            if "날짜" in view_df.columns:
                view_df["날짜"] = pd.to_datetime(view_df["날짜"], errors="coerce")
            if "Block" in view_df.columns:
                view_df["Block"] = pd.to_numeric(view_df["Block"], errors="coerce").fillna(0).astype(int)

            equip_all = inj_equip.copy()
            equip_all["설비명"] = (
                equip_all.get("설비코드", "")
                .astype("string")
                .fillna("")
                .astype(str)
                .str.strip()
                .str.upper()
            )
            equip_all["비고"] = equip_all.get("비고", "").astype("string").fillna("").astype(str).str.strip()
            equip_all["생산 제품"] = equip_all.get("생산 제품", "").astype("string").fillna("").astype(str).str.strip()
            equip_all["배정가능"] = equip_all.get("배정가능", True).fillna(True)

            def _equip_sort_key(s: str) -> tuple[int, int, str]:
                s2 = str(s or "").strip().upper()
                m = re.match(r"^([A-Z])(\d+)$", s2)
                if not m:
                    return (999, 999, s2)
                return (ord(m.group(1)) - 65, int(m.group(2)), s2)

            equip_list = sorted([e for e in equip_all["설비명"].tolist() if str(e).strip()], key=_equip_sort_key)
            equip_info = {
                str(r["설비명"]): {
                    "배정가능": bool(r.get("배정가능", True)),
                    "비고": str(r.get("비고") or "").strip(),
                    "현재제품": str(r.get("생산 제품") or "").strip(),
                }
                for _, r in equip_all.iterrows()
                if str(r.get("설비명") or "").strip()
            }

            slots = []
            for i in range(int(horizon_days)):
                d = start_date + timedelta(days=i)
                for b, sh in [(1, "주간"), (2, "야간")]:
                    slots.append({"날짜": d, "Block": b, "shift": sh, "label": f"{d.month}/{d.day} {sh}"})

            s_map: dict[tuple[str, date, int], dict[str, object]] = {}
            if (not view_df.empty) and all(c in view_df.columns for c in ["설비명", "날짜", "Block"]):
                for _, r in view_df.iterrows():
                    dtv = r.get("날짜")
                    if isinstance(dtv, datetime):
                        dv = dtv.date()
                    elif isinstance(dtv, date):
                        dv = dtv
                    else:
                        continue
                    k = (str(r.get("설비명") or "").strip().upper(), dv, int(r.get("Block") or 0))
                    s_map[k] = r.to_dict()

            cav = _injection_schedule_to_cavity_rows(view_df)
            cav_key: dict[tuple[str, date, int, int], tuple[str, int]] = {}
            if cav is not None and (not cav.empty):
                cav["설비명"] = cav["설비명"].astype("string").fillna("").astype(str).str.strip().str.upper()
                for _, r in cav.iterrows():
                    d = r.get("날짜")
                    if not isinstance(d, date):
                        continue
                    key = (str(r.get("설비명") or "").strip().upper(), d, int(r.get("Block") or 0), int(r.get("CAV") or 0))
                    cav_key[key] = (str(r.get("도수") or "").strip(), int(r.get("필요수량") or 0))

            idx = list(range(1, 9))
            col_blocks: list[tuple[str, str]] = [("", "설비")]
            for sl in slots:
                label = str(sl["label"])
                for sub in ["제품정보", "CAV", "도수", "필요수량"]:
                    col_blocks.append((label, sub))
            cols = pd.MultiIndex.from_tuples(col_blocks, names=["일자/주야", "구분"])

            rows: list[list[object]] = []
            now = datetime.now(KST)
            now_block = 2 if now.hour >= 20 else 1
            # Equipments that have any scheduled assignment in horizon (used to decide fill-running-all).
            sched_assigned: set[str] = set()
            try:
                for (eq, _, _), rec in s_map.items():
                    if str((rec or {}).get("제품명코드") or "").strip():
                        sched_assigned.add(str(eq).strip().upper())
            except Exception:
                sched_assigned = set()
            for equip in equip_list:
                info = equip_info.get(str(equip), {})
                assignable = bool(info.get("배정가능", True))
                note = str(info.get("비고") or "").strip()
                cur_code = str(info.get("현재제품코드") or "").strip().upper()
                cur_name = str(info.get("현재제품") or "").strip()
                fill_running_all = bool(cur_code) and (str(equip).strip().upper() not in sched_assigned)
                for cav_no in idx:
                    row: list[object] = []
                    row.append(str(equip) if cav_no == 1 else "")
                    for sl in slots:
                        d = sl["날짜"]
                        b = int(sl["Block"])
                        rec = s_map.get((str(equip), d, b), {}) if isinstance(d, date) else {}
                        prod = str(rec.get("제품명코드") or "").strip()
                        prod_name = str(rec.get("제품명") or "").strip()
                        if (not prod) and fill_running_all and cur_code:
                            prod = cur_code
                            prod_name = cur_name
                        elif (not prod) and isinstance(d, date) and (d == start_date) and (int(b) == int(now_block)) and cur_code:
                            prod = cur_code
                            prod_name = cur_name
                        prod_info = "\n".join([t for t in [prod, prod_name] if t])
                        if (not prod_info) and note:
                            prod_info = (f"배정불가\n비고: {note}" if (not assignable) else f"유휴\n비고: {note}")
                        if cav_no == 1:
                            row.append(prod_info)
                        else:
                            row.append("")
                        row.append(cav_no)
                        pw, qty = cav_key.get((str(equip), d, b, cav_no), ("", 0))
                        row.append(pw)
                        row.append("" if int(qty) <= 0 else int(qty))
                    rows.append(row)

            op_show = pd.DataFrame(rows, columns=cols)
            st.dataframe(
                _style_dataframe_like_dashboard(op_show),
                use_container_width=True,
                height=_table_height_for_rows(len(op_show), min_height=360, max_height=860),
                hide_index=True,
            )

        if remaining is not None and (not remaining.empty):
            st.divider()
            st.subheader("미배정 잔여수량")
            rem_show = remaining.copy()
            if "잔여수량" in rem_show.columns:
                rem_show["잔여수량"] = pd.to_numeric(rem_show["잔여수량"], errors="coerce").fillna(0).astype(int)
            if "납기일" in rem_show.columns:
                rem_show["납기일"] = pd.to_datetime(rem_show["납기일"], errors="coerce")

            # Fill impacted orders from order-detail (more reliable than demand aggregation which may not carry order keys).
            try:
                if detail_for_map is not None and (not detail_for_map.empty) and ("제품명코드" in rem_show.columns):
                    refs_map = _build_order_refs_by_base_r(detail_for_map)
                    rem_show["영향수주"] = rem_show["제품명코드"].map(lambda k: _format_order_ref_list(refs_map.get(str(k or "").strip().upper(), [])))
            except Exception:
                pass
            sort_cols = [c for c in ["납기일", "잔여수량", "제품명코드"] if c in rem_show.columns]
            if sort_cols:
                asc_map = {"납기일": True, "잔여수량": False, "제품명코드": True}
                rem_show = rem_show.sort_values(
                    sort_cols,
                    ascending=[asc_map.get(c, True) for c in sort_cols],
                    na_position="last",
                )
            xlsx_bytes_rem = _to_excel_bytes(rem_show, sheet_name="미배정")
            st.download_button(
                "엑셀 다운로드 (미배정 잔여)",
                data=xlsx_bytes_rem,
                file_name=f"사출_미배정_{code_label}_{_today_kst().isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"inj_{code_key}_download_rem",
            )
            rem_col_cfg = {}
            if "납기일" in rem_show.columns:
                rem_col_cfg["납기일"] = st.column_config.DatetimeColumn(format="YYYY-MM-DD", width="small")
            if "잔여수량" in rem_show.columns:
                rem_col_cfg["잔여수량"] = st.column_config.NumberColumn(format="localized", width="small")
            if "영향수주" in rem_show.columns:
                rem_col_cfg["영향수주"] = st.column_config.TextColumn(width="large")
            st.dataframe(
                _style_dataframe_like_dashboard(rem_show),
                use_container_width=True,
                height=_table_height_for_rows(len(rem_show), min_height=220, max_height=520),
                hide_index=True,
                column_config=rem_col_cfg if rem_col_cfg else None,
            )
        return

    base_df = df
    if view_mode == "납기별 상세":
        due_quick_state = st.session_state.get("due_due_quick", "해제")
        if due_quick_state != "해제":
            base_df = _apply_due_date_end_filter(base_df, st.session_state.get("due_due_end", _today_kst()))
    if view_mode == "공정별 보기":
        proc_quick_state = st.session_state.get("proc_due_quick", "해제")
        if proc_quick_state != "해제":
            base_df = _apply_due_date_end_filter(base_df, st.session_state.get("proc_due_end", _today_kst()))

    subset = (
        base_df
        if _is_all_codes(codes_selected)
        else base_df[base_df[new_code_col].astype("string").isin(codes_selected)].copy()
    )
    page_key = "due" if process_only is None else f"proc_{process_only}"
    render(
        subset,
        ui_key_prefix=f"{page_key}_{code_key}",
        process_only=process_only,
        selected_code=code_label,
    )


if __name__ == "__main__":
    main()
